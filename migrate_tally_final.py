import openpyxl
from datetime import datetime

# MONKEY-PATCH for openpyxl issue with some Tally exports
from openpyxl.worksheet.filters import FilterColumn
def patched_init(self, colId, hidden=False, customFilters=False, method=None, val=None, **kwargs):
    self.colId, self.hidden, self.customFilters, self.method, self.val = colId, hidden, customFilters, method, val
FilterColumn.__init__ = patched_init

def parse_dt(d):
    if isinstance(d, datetime): return d
    if not d: return None
    for fmt in ('%d-%b-%y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%Y'):
        try: return datetime.strptime(str(d).strip(), fmt)
        except: continue
    return None

def get_account_type(name):
    name = name.lower().strip()
    if any(k in name for k in ['bank', 'cash', 'hdfc', 'kotak', 'indian bank', 'sbi', 'icici']): return 'asset_cash'
    if any(k in name for k in ['receivable', 'debtors', 'railway', 'bill receivable']): return 'asset_receivable'
    if any(k in name for k in ['payable', 'creditors', 'bill payable']): return 'liability_payable'
    if any(k in name for k in ['tax', 'gst', 'igst', 'cgst', 'sgst', 'vat', 'tds']): return 'asset_current'
    if any(k in name for k in ['expense', 'fees', 'salary', 'rent', 'office', 'electricity', 'water', 'repair', 'maintenance', 'travel']): return 'expense'
    if any(k in name for k in ['income', 'sales', 'discount received', 'interest received', 'commission received']): return 'income'
    if any(k in name for k in ['capital', 'equity', 'reserves', 'surplus']): return 'equity'
    if any(k in name for k in ['loan', 'borrowing', 'secured', 'unsecured']): return 'liability_non_current'
    if any(k in name for k in ['fixed asset', 'machinery', 'furniture', 'building', 'land', 'computer', 'vehicle']): return 'asset_fixed'
    return 'asset_current'

def get_or_create_account(name):
    name = str(name).strip()
    if not name or name == 'None' or name == 'NaN': return None
    a = env['account.account'].search([('name', '=', name)], limit=1)
    if not a:
        at = get_account_type(name)
        # Find a code that doesn't exist. Let's use 99xxxx
        last = env['account.account'].search([('code', 'like', '99%')], order='code desc', limit=1)
        nc = str(int(last.code) + 1) if last else '990001'
        a = env['account.account'].create({'name': name, 'code': nc, 'account_type': at})
    return a

def get_or_create_partner(name):
    name = str(name).strip()
    if not name or name == 'None' or name == 'NaN': return None
    p = env['res.partner'].search([('name', '=', name)], limit=1)
    if not p:
        p = env['res.partner'].create({'name': name})
    return p

def migrate():
    fname = '/home/biz/odoo/all_tally_to_odoo_migratation.xlsx'
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    
    j_misc = env['account.journal'].search([('type', '=', 'general')], limit=1)
    j_bank = env['account.journal'].search([('type', '=', 'bank')], limit=1)
    j_sales = env['account.journal'].search([('type', '=', 'sale')], limit=1)
    j_purchase = env['account.journal'].search([('type', '=', 'purchase')], limit=1)
    
    # Map sheets to processing functions
    sheets_to_process = [
        {'name': 'Sales Inv. Register', 'type': 'invoice', 'move_type': 'out_invoice', 'journal': j_sales},
        {'name': 'Sales Inv. Register (2)', 'type': 'invoice', 'move_type': 'out_invoice', 'journal': j_sales},
        {'name': 'Purchase Register', 'type': 'invoice', 'move_type': 'in_invoice', 'journal': j_purchase},
        {'name': 'Debit Note Register', 'type': 'entry', 'journal': j_misc},
        {'name': 'Credit Note Register', 'type': 'entry', 'journal': j_misc},
        {'name': 'Journal Register', 'type': 'entry', 'journal': j_misc},
        {'name': 'Contra Register', 'type': 'entry', 'journal': j_bank},
        {'name': 'Receipt Register', 'type': 'non-columnar', 'journal': j_bank, 'side': 'receipt'},
        {'name': 'Payment Register', 'type': 'non-columnar', 'journal': j_bank, 'side': 'payment'},
    ]
    
    total_created = 0
    
    for s_info in sheets_to_process:
        s_name = s_info['name']
        if s_name not in wb.sheetnames:
            print(f"Skipping {s_name} (not found)")
            continue
            
        print(f"Processing sheet: {s_name}...")
        ws = wb[s_name]
        headers = []
        
        # Determine header row (usually 9, but let's be flexible)
        found_header = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not any(row): continue
            row = [str(cell).strip() if cell is not None else None for cell in row]
            
            if 'Date' in row and ('Particulars' in row or 'Voucher' in row):
                headers = [h if h else f"Col{j}" for j, h in enumerate(row)]
                header_idx = i
                found_header = True
                break
        
        if not found_header:
            print(f"  Could not find headers for {s_name}")
            continue
            
        # Process rows
        for i, row in enumerate(ws.iter_rows(min_row=header_idx + 2, values_only=True)):
            if not any(row): continue
            row_data = dict(zip(headers, row))
            
            dt = parse_dt(row_data.get('Date'))
            if not dt: continue
            
            particulars = str(row_data.get('Particulars') or row_data.get('Particulars (Customer Name)') or '').strip()
            if not particulars or particulars == 'None' or 'Total' in particulars: continue
            
            v_type = str(row_data.get('Voucher Type') or '').strip()
            v_no = str(row_data.get('Voucher No.') or row_data.get('Voucher Ref. No.') or row_data.get('Supplier Invoice No.') or '').strip()
            
            lines = []
            
            if s_info['type'] == 'invoice':
                # For Sales/Purchase, we create lines for each amount column
                partner = get_or_create_partner(particulars)
                
                # Check for existing move to avoid duplicates
                existing = env['account.move'].search([('ref', '=', v_no), ('partner_id', '=', partner.id), ('move_type', '=', s_info['move_type'])], limit=1)
                if existing: continue
                
                for k, v in row_data.items():
                    if any(x in k for x in ['Date', 'Particulars', 'Voucher', 'No.', 'Ref.', 'Gross Total', 'Value', 'Col']): continue
                    if v is None: continue
                    try:
                        val = float(v)
                        if val == 0: continue
                        acc = get_or_create_account(k)
                        if acc:
                            # In Odoo Invoices, we usually create one line per product/tax
                            # For simplicity here, we create lines with accounts
                            lines.append((0, 0, {
                                'account_id': acc.id,
                                'name': k,
                                'debit': val if s_info['move_type'] == 'in_invoice' else 0,
                                'credit': val if s_info['move_type'] == 'out_invoice' else 0,
                            }))
                    except: continue
                
                if not lines: continue
                
                # Odoo will try to balance it with the receivable/payable account automatically if we use 'account.move' with move_type
                # But here we are creating a full entry or an invoice?
                # Let's create an invoice and add lines.
                move = env['account.move'].create({
                    'move_type': s_info['move_type'],
                    'partner_id': partner.id,
                    'date': dt.strftime('%Y-%m-%d'),
                    'invoice_date': dt.strftime('%Y-%m-%d'),
                    'ref': v_no,
                    'journal_id': s_info['journal'].id,
                    'invoice_line_ids': lines if s_info['move_type'] == 'out_invoice' else [], # We use invoice_line_ids for invoices
                })
                # If it's a purchase, we might need a different structure.
                # Actually, for account.move, 'invoice_line_ids' is the way.
                if s_info['move_type'] == 'in_invoice':
                    move.write({'invoice_line_ids': lines})
                
                try:
                    move.action_post()
                    total_created += 1
                except Exception as e:
                    print(f"  ERR Invoicing {v_no}: {e}")
                    move.unlink()
                    
            elif s_info['type'] == 'entry':
                # General Journal, Contra, Debit/Credit Notes (as entries)
                m_acc = get_or_create_account(particulars)
                for k, v in row_data.items():
                    if any(x in k for x in ['Date', 'Particulars', 'Voucher', 'No.', 'Ref.', 'Gross Total', 'Value', 'Col', 'Debit', 'Credit']): continue
                    if v is None: continue
                    try:
                        val = float(v)
                        if val == 0: continue
                        o_acc = get_or_create_account(k)
                        if o_acc:
                            lines.append((0, 0, {
                                'account_id': o_acc.id,
                                'name': f"{v_type} {v_no}",
                                'debit': val if val > 0 else 0,
                                'credit': -val if val < 0 else 0,
                            }))
                    except: continue
                
                # Check for explicit Debit/Credit columns
                dr = float(row_data.get('Debit') or 0)
                cr = float(row_data.get('Credit') or 0)
                if dr != 0: lines.append((0, 0, {'account_id': m_acc.id, 'debit': dr, 'credit': 0, 'name': particulars}))
                if cr != 0: lines.append((0, 0, {'account_id': m_acc.id, 'debit': 0, 'credit': cr, 'name': particulars}))
                
                # Balance it if needed
                bal = sum(l[2]['debit'] for l in lines) - sum(l[2]['credit'] for l in lines)
                if bal != 0:
                    lines.append((0, 0, {
                        'account_id': m_acc.id,
                        'debit': -bal if bal < 0 else 0,
                        'credit': bal if bal > 0 else 0,
                        'name': particulars
                    }))
                
                if not lines: continue
                
                move = env['account.move'].create({
                    'move_type': 'entry',
                    'date': dt.strftime('%Y-%m-%d'),
                    'ref': v_no,
                    'journal_id': s_info['journal'].id,
                    'line_ids': lines,
                })
                try:
                    move.action_post()
                    total_created += 1
                except Exception as e:
                    print(f"  ERR Entry {v_no}: {e}")
                    
            elif s_info['type'] == 'non-columnar':
                # Receipt / Payment
                partner = get_or_create_partner(particulars)
                try: 
                    dr = float(row_data.get('Debit') or 0)
                    cr = float(row_data.get('Credit') or 0)
                except: dr, cr = 0, 0
                
                if dr == 0 and cr == 0: continue
                
                # For receipts/payments, we often have a bank account on the other side
                # If it's a receipt, debit bank, credit partner (receivable or income)
                # If it's a payment, credit bank, debit partner (payable or expense)
                
                bank_acc = s_info['journal'].default_account_id
                part_acc = get_or_create_account(particulars) # This might be the partner's account or a direct expense/income
                
                if s_info['side'] == 'receipt':
                    # Debit Bank, Credit Part
                    lines = [
                        (0, 0, {'account_id': bank_acc.id, 'debit': cr if cr > 0 else dr, 'credit': 0, 'name': f"Receipt {v_no}"}),
                        (0, 0, {'account_id': part_acc.id, 'debit': 0, 'credit': cr if cr > 0 else dr, 'name': particulars}),
                    ]
                else: # Payment
                    # Credit Bank, Debit Part
                    lines = [
                        (0, 0, {'account_id': bank_acc.id, 'debit': 0, 'credit': dr if dr > 0 else cr, 'name': f"Payment {v_no}"}),
                        (0, 0, {'account_id': part_acc.id, 'debit': dr if dr > 0 else cr, 'credit': 0, 'name': particulars}),
                    ]
                
                move = env['account.move'].create({
                    'move_type': 'entry',
                    'date': dt.strftime('%Y-%m-%d'),
                    'ref': v_no,
                    'journal_id': s_info['journal'].id,
                    'line_ids': lines,
                })
                try:
                    move.action_post()
                    total_created += 1
                except Exception as e:
                    print(f"  ERR Pay/Rec {v_no}: {e}")

            if total_created % 100 == 0:
                env.cr.commit()
                print(f"  Processed {total_created} records...")

    env.cr.commit()
    print(f"Migration complete. Total records: {total_created}")

migrate()
