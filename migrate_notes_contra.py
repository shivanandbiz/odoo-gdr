# migrate_notes_contra.py
import openpyxl
from datetime import datetime
from collections import defaultdict

def get_partner(name):
    p = env['res.partner'].search([('name', '=', name)], limit=1)
    if not p:
        p = env['res.partner'].create({'name': name})
    return p.id

def get_account(name):
    a = env['account.account'].search([('name', '=', name)], limit=1)
    if not a:
        code = 'MIG' + name[:5].upper()
        a = env['account.account'].create({'name': name, 'code': code, 'account_type': 'asset_current'})
    return a.id

def get_tax(name, amount, type):
    t = env['account.tax'].search([('name', '=', name), ('type_tax_use', '=', type)], limit=1)
    if not t:
        t = env['account.tax'].create({'name': name, 'amount': amount, 'type_tax_use': type, 'amount_type': 'percent'})
    return t.id

def fval(d, *keys):
    for k in keys:
        try:
            v = d.get(k)
            if v is not None: return float(v)
        except: pass
    return 0.0

def migrate_contra():
    print("Migrating Contra Register...")
    wb = openpyxl.load_workbook('/home/biz/odoo/all_tally_to_odoo_migratation.xlsx', read_only=True, data_only=True)
    ws = wb['Contra Register']
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = 8
    headers = [str(c).strip() if c is not None else f'Col{j}' for j, c in enumerate(rows[hdr_idx])]
    
    # Ledger cols from index 4 onwards
    ledger_cols = []
    for j in range(4, len(headers)):
        if headers[j] and 'Unnamed' not in headers[j]:
            ledger_cols.append((j, headers[j]))
            
    journal = env['account.journal'].search([('type', '=', 'general')], limit=1)
    count = 0
    for row in rows[hdr_idx + 1:]:
        d = {h: v for h, v in zip(headers, row)}
        part = str(d.get('Particulars', '') or '').strip()
        if not part or 'Total' in part or 'Grand' in part: continue
        
        dt = d.get('Date')
        if not isinstance(dt, datetime): continue
        
        lines = []
        sum_debits = 0
        for col_idx, lname in ledger_cols:
            val = float(str(row[col_idx] or 0).strip() or 0)
            if val != 0:
                lines.append({
                    'name': lname,
                    'account_id': get_account(lname),
                    'debit': val if val > 0 else 0,
                    'credit': -val if val < 0 else 0
                })
                sum_debits += val
        
        if not lines: continue
        
        # Balance with Particulars
        lines.append({
            'name': part,
            'account_id': get_account(part),
            'debit': 0 if sum_debits > 0 else -sum_debits,
            'credit': sum_debits if sum_debits > 0 else 0
        })
        
        env['account.move'].create({
            'move_type': 'entry',
            'journal_id': journal.id,
            'date': dt.strftime('%Y-%m-%d'),
            'line_ids': [(0, 0, ln) for ln in lines],
        }).action_post()
        count += 1
    env.cr.commit()
    print(f"Finished Contra: {count} records.")

def migrate_notes(sheet_name, move_type, tax_use):
    print(f"Migrating {sheet_name}...")
    wb = openpyxl.load_workbook('/home/biz/odoo/all_tally_to_odoo_migratation.xlsx', read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = 8
    headers = [str(c).strip() if c is not None else f'Col{j}' for j, c in enumerate(rows[hdr_idx])]
    
    count = 0
    for row in rows[hdr_idx + 1:]:
        d = {h: v for h, v in zip(headers, row)}
        part = str(d.get('Particulars', '') or '').strip()
        if not part or 'Total' in part or 'Grand' in part: continue
        
        dt = d.get('Date')
        if not isinstance(dt, datetime): continue
        
        gross = float(str(d.get('Gross Total') or 0).strip() or 0)
        
        igst18 = fval(d, 'IGST@18', 'IGST@ 18')
        cgst9 = fval(d, 'CGST @ 9%', 'CGST@9%')
        sgst9 = fval(d, 'SGST@ 9%', 'SGST@9%')
        total_gst = igst18 + cgst9 + sgst9
        taxable_base = gross - total_gst
        
        tax_ids = []
        if igst18 > 0: tax_ids = [get_tax('IGST 18%', 18.0, tax_use)]
        elif cgst9 > 0 or sgst9 > 0: 
            tax_ids = [get_tax(f'CGST 9%', 9.0, tax_use), get_tax(f'SGST 9%', 9.0, tax_use)]
            
        env['account.move'].create({
            'move_type': move_type,
            'partner_id': get_partner(part),
            'invoice_date': dt.strftime('%Y-%m-%d'),
            'date': dt.strftime('%Y-%m-%d'),
            'ref': str(d.get('Voucher Ref. No.') or ''),
            'invoice_line_ids': [(0, 0, {
                'name': 'Note Migration',
                'quantity': 1,
                'price_unit': round(taxable_base, 2),
                'tax_ids': [(6, 0, tax_ids)],
            })],
        }).action_post()
        count += 1
    env.cr.commit()
    print(f"Finished {sheet_name}: {count} records.")

migrate_contra()
migrate_notes('Debit Note Register', 'in_refund', 'purchase')
migrate_notes('Credit Note Register', 'out_refund', 'sale')
