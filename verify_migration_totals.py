import openpyxl
from datetime import datetime
from collections import defaultdict

def read_purchase_sheet(fname, sheet_name='Purchase Register'):
    print(f"Reading {fname} [{sheet_name}]...")
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  Error: {sheet_name} not in {fname}")
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    hdr_idx = None
    for i, row in enumerate(rows):
        if any(str(c).strip() == 'Date' for c in row if c is not None):
            hdr_idx = i
            break
    if hdr_idx is None:
        print(f"  Error: No header found in {fname}")
        return []
    
    headers = [str(c).strip() if c is not None else f'Col{j}' for j, c in enumerate(rows[hdr_idx])]
    data = []
    for row in rows[hdr_idx + 1:]:
        d = {h: v for h, v in zip(headers, row)}
        part = str(d.get('Particulars', '') or '').strip()
        if not part or 'Total' in part or 'Grand' in part or part == 'nan':
            continue
        
        raw_date = d.get('Date')
        if raw_date is None or not isinstance(raw_date, (datetime, str)):
            continue
        
        try:
            if isinstance(raw_date, datetime):
                dt = raw_date
            else:
                dt = datetime.strptime(str(raw_date).strip()[:10], '%Y-%m-%d')
        except:
            continue
            
        d['_date'] = dt
        # Handle shifted columns or non-numeric Gross Total
        gross = d.get('Gross Total')
        try:
            d['_gross'] = float(gross or 0)
        except:
            # If gross is NOT a float (like a date in March File 1), 
            # we might need to look at 'LOCAL PURCHASE GST 18%' which appeared to hold the gross
            try:
                # Try reading from LOCAL PURCHASE GST 18% if it looks like a number
                alt_gross = d.get('LOCAL PURCHASE GST 18%')
                d['_gross'] = float(alt_gross or 0)
                # print(f"  Note: Using alt gross for row {dt} {part}: {d['_gross']}")
            except:
                d['_gross'] = 0.0
        
        data.append(d)
    return data

# Source files
file1 = '/home/biz/odoo/all_tally_to_odoo_migratation.xlsx'
file2 = '/home/biz/odoo/purchase.xlsx-2025-26.......xlsx'

rows1 = read_purchase_sheet(file1)
rows2 = read_purchase_sheet(file2)

# Merge logic
# April - December 2025: File 1
# January 2026: File 1 (We need to investigate this one)
# February - March 2026: File 2

final_rows = []
monthly_check = defaultdict(float)

for r in rows1:
    dt = r['_date']
    my = dt.strftime('%Y-%m')
    if my in ['2025-04', '2025-05', '2025-06', '2025-07', '2025-08', '2025-09', '2025-10', '2025-11', '2025-12', '2026-01']:
        final_rows.append(r)
        monthly_check[my] += r['_gross']

for r in rows2:
    dt = r['_date']
    my = dt.strftime('%Y-%m')
    if my in ['2026-02', '2026-03']:
        final_rows.append(r)
        monthly_check[my] += r['_gross']

print("\nResulting Monthly Totals:")
target = {
    '2025-04': 4965142.20,
    '2025-05': 5935852.36,
    '2025-06': 5902935.38,
    '2025-07': 5832890.98,
    '2025-08': 2706354.30,
    '2025-09': 3742037.02,
    '2025-10': 5164669.18,
    '2025-11': 4396926.92,
    '2025-12': 4792170.86,
    '2026-01': 7199291.74,
    '2026-02': 9630540.24,
    '2026-03': 30054329.14
}

for m in sorted(target.keys()):
    actual = monthly_check.get(m, 0)
    diff = actual - target[m]
    print(f"{m}: Actual={actual:12.2f} | Target={target[m]:12.2f} | Diff={diff:12.2f}")

print(f"\nTotal Rows: {len(final_rows)}")
print(f"Grand Total: {sum(monthly_check.values()):,.2f}")
print(f"Target Grand: 90,323,140.32")
