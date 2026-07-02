import openpyxl

wb = openpyxl.load_workbook('/home/biz/odoo/all_tally_to_odoo_migratation.xlsx', read_only=True, data_only=True)
ws = wb['Receipt Register']

# Get headers from Row 9
headers = [cell.value for cell in next(ws.iter_rows(min_row=9, max_row=9))]
print(f"Headers found: {headers}")

for i, row in enumerate(ws.iter_rows(min_row=11, max_row=50, values_only=True)):
    print(f"Row {i+11}: {row}")
