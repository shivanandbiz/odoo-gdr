# migrate_payments_v5.py
import openpyxl
from datetime import datetime

# PATCH openpyxl for Tally's invalid XML filters
from openpyxl.worksheet.filters import FilterColumn
def patched_init(self, colId, hidden=False, customFilters=False, method=None, val=None, **kwargs):
    self.colId, self.hidden, self.customFilters, self.method, self.val = colId, hidden, customFilters, method, val
FilterColumn.__init__ = patched_init

def get_account_type(name):
    name = name.lower().strip()
    if any(k in name for k in ['bank', 'cash', 'hdfc', 'kotak', 'indian bank', 'axis', 'sbi', 'icici']): return 'asset_cash'
    if any(k in name for k in ['receivable', 'debtors']): return 'asset_receivable'
    if any(k in name for k in ['payable', 'creditors', 'gst payable']): return 'liability_payable'
    if any(k in name for k in ['tax', 'gst', 'igst', 'cgst', 'sgst']): return 'asset_current'
    if any(k in name for k in ['income', 'sales']): return 'income'
    if any(k in name for k in ['expense', 'purchase', 'salary', 'rent']): return 'expense'
    return 'asset_current'

def get_or_create_account(name):
    name = str(name).strip()
    if not name or name == 'None' or name == '': return None
    a = env['account.account'].search([('name', '=', name)], limit=1)
    if not a:
        at = get_account_type(name)
        last = env['account.account'].search([('code', 'like', '99%')], order='code desc', limit=1)
        nc = str(int(last.code) + 1) if last else '990001'
        a = env['account.account'].create({'name': name, 'code': nc, 'account_type': at})
    return a

def find_partner(name):
    p = env['res.partner'].search([('name', '=', name)], limit=1)
    return p.id if p else None

def parse_dt(d):
    if isinstance(d, datetime): return d
    if not d: return None
    s = str(d).strip()
    if not s or s == 'None': return None
    for fmt in ('%d-%b-%y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%Y'):
        try: return datetime.strptime(s, fmt)
        except: continue
    return None

def create_move(dt, vtype, vno, lines, journal_id):
    if not lines: return
    # Balance check
    bal = sum(l['debit'] for l in lines) - sum(l['credit'] for l in lines)
    if abs(bal) > 0.01:
        # Try to find if one of the lines was a bank/cash account to adjust
        # If not, add balancing line to default account of journal
        journal = env['account.journal'].browse(journal_id)
        lines.append({
            'account_id': journal.default_account_id.id or get_or_create_account('Suspense Account').id,
            'debit': -bal if bal < 0 else 0,
            'credit': bal if bal > 0 else 0,
            'name': f"Balancing {vtype} {vno}"
        })
    
    vals = {
        'move_type': 'entry',
        'date': dt.strftime('%Y-%m-%d') if dt else datetime.now().strftime('%Y-%m-%d'),
        'ref': f"{vtype} {vno}" if vtype and vno else (vno or vtype or ''),
        'journal_id': journal_id,
        'line_ids': [(0,0, l) for l in lines]
    }
    try:
        move = env['account.move'].create(vals)
        move.action_post()
        return move
    except Exception as e:
        print(f"  ERR Create {vno}: {e}")
        return None

def migrate():
    fname = '/home/biz/odoo/all_tally_to_odoo_migratation.xlsx'
    j_bank = env['account.journal'].search([('type', '=', 'bank')], limit=1).id
    j_misc = env['account.journal'].search([('type', '=', 'general')], limit=1).id
    j_sale = env['account.journal'].search([('type', '=', 'sale')], limit=1).id
    j_purch = env['account.journal'].search([('type', '=', 'purchase')], limit=1).id

    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    
    registers = [
        {'name': 'Contra Register', 'jid': j_bank},
        {'name': 'Receipt Register', 'jid': j_bank},
        {'name': 'Payment Register', 'jid': j_bank},
        {'name': 'Journal Register', 'jid': j_misc},
        {'name': 'Debit Note Register', 'jid': j_purch},
        {'name': 'Credit Note Register', 'jid': j_sale},
    ]

    for reg in registers:
        print(f"Processing {reg['name']}...")
        if reg['name'] not in wb.sheetnames:
            print(f"  Sheet {reg['name']} not found.")
            continue
        ws = wb[reg['name']]
        
        last_date = None
        last_vtype = None
        last_vno = None
        current_lines = []
        tc = 0
        
        # Find header dynamically
        header_row = -1
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if 'Vch No.' in [str(x) for x in row]:
                header_row = i
                break
        
        if header_row == -1:
            print(f"  Header not found in {reg['name']}")
            continue

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_row: continue
            if not any(row): continue
            
            p_name = str(row[1]).strip() if row[1] else ''
            if not p_name or 'Total' in p_name: continue
            
            dt = parse_dt(row[0])
            vtype = str(row[2]).strip() if row[2] else None
            vno = str(row[3]).strip() if row[3] else None
            
            # Carry forward
            if not dt: dt = last_date
            if not vtype: vtype = last_vtype
            if not vno: vno = last_vno

            # If this is a NEW voucher starting (based on Date/Type/No change)
            # Actually in Tally registers, if Date and Vch No are present, it's a new voucher.
            is_new = False
            if row[0] is not None or row[3] is not None:
                is_new = True

            if is_new and current_lines:
                create_move(last_date, last_vtype, last_vno, current_lines, reg['jid'])
                tc += 1
                current_lines = []
            
            last_date = dt
            last_vtype = vtype
            last_vno = vno
            
            try:
                dr = float(row[4] or 0)
                cr = float(row[5] or 0)
            except: dr, cr = 0, 0
            
            if dr == 0 and cr == 0: continue
            
            acc = get_or_create_account(p_name)
            if acc:
                current_lines.append({
                    'account_id': acc.id,
                    'debit': dr,
                    'credit': cr,
                    'name': f"{vtype} {vno}" if vtype and vno else p_name,
                    'partner_id': find_partner(p_name)
                })

        # Final one
        if current_lines:
            create_move(last_date, last_vtype, last_vno, current_lines, reg['jid'])
            tc += 1
        
        print(f"  Created {tc} items from {reg['name']}")
        env.cr.commit()

migrate()
