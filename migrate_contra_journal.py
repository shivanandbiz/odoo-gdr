# migrate_contra_journal.py
import openpyxl
from datetime import datetime

from openpyxl.worksheet.filters import FilterColumn, CustomFilter
def patched_init(self, colId, hidden=False, customFilters=False, method=None, val=None, **kwargs):
    self.colId, self.hidden, self.customFilters, self.method, self.val = colId, hidden, customFilters, method, val
FilterColumn.__init__ = patched_init

def patched_custom_filter_init(self, operator=None, val=None, **kwargs):
    self.operator = operator
    self.__dict__['val'] = val
CustomFilter.__init__ = patched_custom_filter_init


def get_account_type(name):
    name = name.lower().strip()
    if any(k in name for k in ['bank', 'cash', 'hdfc', 'kotak', 'indian bank', 'axis', 'sbi', 'icici']): return 'asset_cash'
    if any(k in name for k in ['receivable', 'debtors']): return 'asset_receivable'
    if any(k in name for k in ['payable', 'creditors', 'gst payable']): return 'liability_payable'
    if any(k in name for k in ['tax', 'gst', 'igst', 'cgst', 'sgst']): return 'asset_current'
    if any(k in name for k in ['income', 'sales', 'discount']): return 'income'
    if any(k in name for k in ['expense', 'purchase', 'salary', 'rent', 'fees', 'charges']): return 'expense'
    if any(k in name for k in ['capital', 'equity']): return 'equity'
    if any(k in name for k in ['loan']): return 'liability_non_current'
    return 'asset_current'

NEXT_CODE = [5000000]
account_cache = {}

def get_or_create_account(name):
    name = str(name).strip()
    if not name or name == 'None' or name == '': return None
    if name in account_cache:
        return account_cache[name]
    a = env['account.account'].search([('name', '=', name)], limit=1)
    if not a:
        at = get_account_type(name)
        NEXT_CODE[0] += 1
        nc = str(NEXT_CODE[0])
        a = env['account.account'].create({'name': name, 'code': nc, 'account_type': at})
    account_cache[name] = a
    return a


def find_partner(name):
    p = env['res.partner'].search([('name', '=', name)], limit=1)
    return p.id if p else None

def parse_dt(d):
    if isinstance(d, datetime): return d
    if not d: return None
    s = str(d).strip()
    if not s or s == 'None': return None
    for fmt in ('%d-%b-%y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%Y'):
        try: return datetime.strptime(s, fmt)
        except: continue
    return None

def create_move(dt, vtype, vno, lines, jid):
    if not lines: return
    bal = sum(l['debit'] for l in lines) - sum(l['credit'] for l in lines)
    if abs(bal) > 0.01:
        j = env['account.journal'].browse(jid)
        lines.append({
            'account_id': j.default_account_id.id or get_or_create_account('Suspense Account').id,
            'debit': -bal if bal < 0 else 0,
            'credit': bal if bal > 0 else 0,
            'name': f"Balancing {vtype} {vno}"
        })
    vals = {
        'move_type': 'entry',
        'date': dt.strftime('%Y-%m-%d') if dt else datetime.now().strftime('%Y-%m-%d'),
        'ref': f"{vtype} {vno}" if vtype and vno else (vno or vtype or ''),
        'journal_id': jid,
        'line_ids': [(0,0, l) for l in lines]
    }
    try:
        m = env['account.move'].create(vals)
        m.action_post()
        return m
    except Exception as e:
        print(f"  ERR Create {vno}: {e}")
        return None

def migrate():
    fname = '/home/biz/odoo/all_tally_to_odoo_migratation.xlsx'
    j_bank = env['account.journal'].search([('type', '=', 'bank')], limit=1).id
    j_misc = env['account.journal'].search([('type', '=', 'general')], limit=1).id

    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    
    registers = [
        {'name': 'Contra Register', 'jid': j_bank},
        {'name': 'Journal Register', 'jid': j_misc},
    ]

    for reg in registers:
        print(f"Processing {reg['name']}...")
        if reg['name'] not in wb.sheetnames:
            print(f"  Sheet {reg['name']} not found.")
            continue
        ws = wb[reg['name']]
        
        last_date = None
        last_vtype = None
        last_vno = None
        current_lines = []
        tc = 0
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= 8: continue
            if not any(row): continue
            
            p_name = str(row[1]).strip() if row[1] else ''
            if not p_name or 'Total' in p_name: continue
            
            dt = parse_dt(row[0])
            vtype = str(row[2]).strip() if row[2] else None
            vno = str(row[3]).strip() if row[3] else None
            
            is_new = False
            if row[0] is not None or row[3] is not None:
                is_new = True

            if is_new and current_lines:
                create_move(last_date, last_vtype, last_vno, current_lines, reg['jid'])
                tc += 1
                current_lines = []
            
            if dt: last_date = dt
            if vtype: last_vtype = vtype
            if vno: last_vno = vno
            
            try:
                dr = float(row[4] or 0)
                cr = float(row[5] or 0)
            except: dr, cr = 0, 0
            
            if dr == 0 and cr == 0: continue
            
            acc = get_or_create_account(p_name)
            if acc:
                current_lines.append({
                    'account_id': acc.id,
                    'debit': dr,
                    'credit': cr,
                    'name': p_name,
                    'partner_id': find_partner(p_name)
                })

        # Final one
        if current_lines:
            create_move(last_date, last_vtype, last_vno, current_lines, reg['jid'])
            tc += 1
        
        print(f"  Created {tc} items from {reg['name']}")
        env.cr.commit()

migrate()
