# migrate_notes.py
import openpyxl
from datetime import datetime

from openpyxl.worksheet.filters import FilterColumn, CustomFilter
def patched_init(self, colId, hidden=False, customFilters=False, method=None, val=None, **kwargs):
    self.colId, self.hidden, self.customFilters, self.method, self.val = colId, hidden, customFilters, method, val
FilterColumn.__init__ = patched_init

def patched_custom_filter_init(self, operator=None, val=None, **kwargs):
    self.operator = operator
    self.__dict__['val'] = val
CustomFilter.__init__ = patched_custom_filter_init


def get_account_type(name):
    name = name.lower().strip()
    if any(k in name for k in ['bank', 'cash', 'hdfc', 'kotak', 'indian bank', 'axis', 'sbi', 'icici']): return 'asset_cash'
    if any(k in name for k in ['receivable', 'debtors', 'railway']): return 'asset_receivable'
    if any(k in name for k in ['payable', 'creditors']): return 'liability_payable'
    if any(k in name for k in ['tax', 'gst', 'igst', 'cgst', 'sgst']): return 'asset_current'
    if any(k in name for k in ['income', 'sales', 'discount']): return 'income'
    if any(k in name for k in ['expense', 'purchase', 'salary', 'rent', 'fees', 'charges']): return 'expense'
    if any(k in name for k in ['capital', 'equity']): return 'equity'
    if any(k in name for k in ['loan']): return 'liability_non_current'
    return 'asset_current'

NEXT_CODE = [5000000]
account_cache = {}

def get_or_create_account(name):
    name = str(name).strip()
    if not name or name == 'None' or name == '': return None
    if name in account_cache:
        return account_cache[name]
    a = env['account.account'].search([('name', '=', name)], limit=1)
    if not a:
        at = get_account_type(name)
        NEXT_CODE[0] += 1
        nc = str(NEXT_CODE[0])
        a = env['account.account'].create({'name': name, 'code': nc, 'account_type': at})
    account_cache[name] = a
    return a

def find_partner(name):
    p = env['res.partner'].search([('name', '=', name)], limit=1)
    if not p:
        p = env['res.partner'].create({'name': name})
    return p

def parse_dt(d):
    if isinstance(d, datetime): return d
    if not d: return None
    s = str(d).strip()
    if not s or s == 'None': return None
    for fmt in ('%d-%b-%y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%Y'):
        try: return datetime.strptime(s, fmt)
        except: continue
    return None

def migrate():
    fname = '/home/biz/odoo/all_tally_to_odoo_migratation.xlsx'
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    
    registers = [
        {'name': 'Debit Note Register', 'move_type': 'in_refund'}, # Vendor Credit Note (Debit Note in Tally)
        {'name': 'Credit Note Register', 'move_type': 'out_refund'} # Customer Credit Note
    ]

    for reg in registers:
        print(f"\nProcessing {reg['name']}...")
        if reg['name'] not in wb.sheetnames:
            print(f"  Sheet {reg['name']} not found.")
            continue
        ws = wb[reg['name']]
        
        headers = []
        tc = 0
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 8:
                headers = [str(x).strip() if x else f"Col{j}" for j, x in enumerate(row)]
                continue
            if i < 9: continue
            
            p_name = str(row[1]).strip() if row[1] else ''
            if not p_name or 'Total' in p_name or p_name == 'None': continue
            
            dt = parse_dt(row[0])
            if not dt: continue
            
            vtype = str(row[2]).strip()
            vno = str(row[3]).strip()
            
            partner = find_partner(p_name)
            
            invoice_vals = {
                'move_type': reg['move_type'],
                'partner_id': partner.id,
                'invoice_date': dt.strftime('%Y-%m-%d'),
                'date': dt.strftime('%Y-%m-%d'),
                'ref': f"{vtype} {vno}",
                'invoice_line_ids': []
            }
            
            lines = []
            
            # Map columns to items/taxes
            for j in range(7, len(row)):
                if j >= len(headers): break
                col_name = headers[j]
                val = row[j]
                try: val = float(val)
                except: val = 0.0
                
                if val == 0: continue
                
                if any(x in col_name.lower() for x in ['tax', 'gst', 'igst', 'cgst', 'sgst']):
                    # Ensure tax account exists
                    tax_acc = get_or_create_account(col_name)
                    # We add tax as a regular accounting line for now, to exactly match balance
                    # Because standard invoice_line_ids with taxes might compute their own values.
                    # Or we can just use move_type='entry' for notes to be 100% safe.
                    # Actually, if we use 'entry', we must provide debit/credit lines.
                    # Let's add it as a normal invoice line for simplicity, without odoo calculating taxes natively.
                    lines.append((0, 0, {
                        'name': col_name,
                        'account_id': tax_acc.id,
                        'quantity': 1,
                        'price_unit': val,
                    }))
                else:
                    # Normal income/expense
                    acc = get_or_create_account(col_name)
                    lines.append((0, 0, {
                        'name': col_name,
                        'account_id': acc.id,
                        'quantity': 1,
                        'price_unit': val,
                    }))
            
            if lines:
                invoice_vals['invoice_line_ids'] = lines
                try:
                    move = env['account.move'].create(invoice_vals)
                    move.action_post()
                    tc += 1
                except Exception as e:
                    print(f"  ERR {vno}: {e}")
                    
        print(f"  Created {tc} notes for {reg['name']}")
        env.cr.commit()

migrate()
