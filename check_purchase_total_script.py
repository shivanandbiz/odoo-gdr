import openpyxl

from openpyxl.worksheet.filters import FilterColumn, CustomFilter
def patched_init(self, colId, hidden=False, customFilters=False, method=None, val=None, **kwargs):
    self.colId, self.hidden, self.customFilters, self.method, self.val = colId, hidden, customFilters, method, val
FilterColumn.__init__ = patched_init

def patched_custom_filter_init(self, operator=None, val=None, **kwargs):
    self.operator = operator
    self.__dict__['val'] = val
CustomFilter.__init__ = patched_custom_filter_init

def check_file(fname):
    print(f"Checking {fname}")
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    if 'Purchase Register' not in wb.sheetnames:
        print("No Purchase Register")
        return
    ws = wb['Purchase Register']
    
    header_row = -1
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        str_row = [str(x) for x in row if x]
        if any('Vch No' in x for x in str_row):
            header_row = i
            break
            
    if header_row == -1: return
    print(f"Found headers at row {header_row}")
    
    total_val = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row: continue
        if not row[1] or 'Total' in str(row[1]): continue
        
        # Typically in purchase register from Tally, amount is row 4 or Gross Total is further down.
        # Let's print the first row after header to debug
        if i == header_row + 1:
            print("First row:", row)
        
        # In tally purchase register, amount is usually at index 4 (Value) and Gross Total is 5
        # Let's try to sum up idx 5 if available
        val = row[5] if len(row) > 5 else row[4]
        try:
            total_val += float(val)
        except:
            pass
            
    print(f"File Total: {total_val:,.2f}")

check_file('/home/biz/odoo/purchase.xlsx-2025-26.......xlsx')
