import openpyxl
from datetime import datetime
import psycopg2

def parse_dt(d):
    if isinstance(d, datetime): return d
    if not d: return None
    for fmt in ('%d-%b-%y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%Y'):
        try: return datetime.strptime(str(d).strip(), fmt)
        except: continue
    return None

def run_buggy_restore():
    db_params = {'dbname': 'Odoo', 'user': 'odoo', 'password': 'odoo', 'host': 'localhost'}
    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()

    # Delete current Feb bills
    cur.execute("SELECT id FROM account_move WHERE move_type = 'in_invoice' AND TO_CHAR(date, 'YYYY-MM') = '2026-02'")
    move_ids = [r[0] for r in cur.fetchall()]
    if move_ids:
        cur.execute("DELETE FROM account_move_line WHERE move_id IN %s", (tuple(move_ids),))
        cur.execute("DELETE FROM account_move WHERE id IN %s", (tuple(move_ids),))

    fname = '/home/biz/odoo/all_tally_to_odoo_migratation.xlsx'
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    ws = wb['Purchase Register']
    
    headers = []
    header_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if 'Date' in [str(c).strip() for c in row if c is not None]:
            headers = [str(h).strip() if h else f'Col{j}' for j, h in enumerate(row)]
            header_idx = i
            break
            
    total_created = 0
    for i, row in enumerate(ws.iter_rows(min_row=header_idx + 2, values_only=True)):
        row_data = dict(zip(headers, row))
        dt = parse_dt(row_data.get('Date'))
        if not dt or dt.month != 2: continue
        
        particulars = str(row_data.get('Particulars') or '').strip()
        if not particulars or 'Total' in particulars: continue
        
        v_ref = str(row_data.get('Supplier Invoice No.') or '').strip()
        
        cur.execute("SELECT id FROM res_partner WHERE name = %s LIMIT 1", (particulars,))
        p = cur.fetchone()
        pid = p[0] if p else 1
        
        try: amount_total = float(row_data.get('Gross Total') or 0)
        except: amount_total = 0

        # INSERT BUGGY MOVE
        cur.execute("""
            INSERT INTO account_move (name, ref, date, invoice_date, partner_id, move_type, state, journal_id, company_id, currency_id, amount_total, auto_post)
            VALUES ('/', %s, %s, %s, %s, 'in_invoice', 'posted', 2, 1, 3, %s, 'no') RETURNING id
        """, (v_ref, dt.strftime('%Y-%m-%d'), dt.strftime('%Y-%m-%d'), pid, amount_total))
        mid = cur.fetchone()[0]
        
        # INSERT PAYABLE (Credit)
        cur.execute("""
            INSERT INTO account_move_line (move_id, account_id, partner_id, name, debit, credit, date, company_id, balance, currency_id, display_type)
            VALUES (%s, 15, %s, %s, 0, %s, %s, 1, %s, 3, 'payment_term')
        """, (mid, pid, v_ref, amount_total, dt.strftime('%Y-%m-%d'), -amount_total))

        # INSERT EXPENSE/TAX (Debit) - THE BUGGY PART
        for k, v in row_data.items():
            if any(x in k for x in ['Date', 'Particulars', 'Voucher', 'No.', 'Ref.', 'Gross Total', 'Value', 'Col']): continue
            if v is None: continue
            try:
                val = float(v)
                if val == 0: continue
                cur.execute("SELECT id FROM account_account WHERE name::text ILIKE %s LIMIT 1", (f'%{k}%',))
                a = cur.fetchone()
                aid = a[0] if a else 29
                
                cur.execute("""
                    INSERT INTO account_move_line (move_id, account_id, partner_id, name, debit, credit, date, company_id, balance, currency_id, display_type)
                    VALUES (%s, %s, %s, %s, %s, 0, %s, 1, %s, 3, 'product')
                """, (mid, aid, pid, k, val, dt.strftime('%Y-%m-%d'), val))
            except: continue
            
        total_created += 1
    
    conn.commit()
    print(f'Restored {total_created} buggy records.')

if __name__ == "__main__":
    run_buggy_restore()
