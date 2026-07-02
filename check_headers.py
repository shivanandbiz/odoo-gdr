import openpyxl
from openpyxl.worksheet.filters import FilterColumn
def patched_init(self, colId, hidden=False, customFilters=False, method=None, val=None, **kwargs):
    self.colId, self.hidden, self.customFilters, self.method, self.val = colId, hidden, customFilters, method, val
FilterColumn.__init__ = patched_init

wb = openpyxl.load_workbook('/home/biz/odoo/all_tally_to_odoo_migratation.xlsx', read_only=True, data_only=True)
sheets = ['Contra Register', 'Journal Register', 'Debit Note Register', 'Credit Note Register']
for sh in sheets:
    if sh not in wb.sheetnames:
        print(f"Sheet {sh} not found.")
        continue
    ws = wb[sh]
    print(f"\n--- {sh} ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if 8 <= i <= 15:
            # clean up None
            clean = [str(x) if x is not None else '' for x in row[:15]]
            print(f"Row {i}: {clean}")
        if i > 15:
            break
