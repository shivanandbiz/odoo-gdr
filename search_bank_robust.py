import openpyxl

file_path = '/home/biz/odoo/all_tally_to_odoo_migratation.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

# Try each sheet independently to isolate the error
for sheet_name in wb.sheetnames:
    print(f"Searching sheet: {sheet_name}")
    try:
        ws = wb[sheet_name]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            for c_idx, cell in enumerate(row):
                if cell and ("HDFC" in str(cell) or "Kotak" in str(cell) or "Indian Bank" in str(cell)):
                    print(f"  FOUND: Row: {r_idx+1}, Col: {c_idx+1}, Value: {cell}")
    except Exception as e:
        print(f"  Error on sheet {sheet_name}: {e}")
