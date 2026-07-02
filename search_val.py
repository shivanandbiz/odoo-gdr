import openpyxl

wb = openpyxl.load_workbook('/home/biz/odoo/all_tally_to_odoo_migratation.xlsx', read_only=True, data_only=True)

target_val = 1023060 # From Receipt 1
for sheet_name in wb.sheetnames:
    print(f"Searching sheet: {sheet_name}")
    try:
        ws = wb[sheet_name]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            for c_idx, cell in enumerate(row):
                if cell == target_val:
                    print(f"  MATCH FOUND: Row: {r_idx+1}, Col: {c_idx+1}, Sheet: {sheet_name}")
    except:
        pass
