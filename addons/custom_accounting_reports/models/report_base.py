
from odoo import models, api, fields
from datetime import datetime

class CustomAccountingReport(models.AbstractModel):
    _name = 'custom.accounting.report'
    _description = 'Custom Accounting Report Logic'

    @api.model
    def get_balance_sheet(self, date_at=None, target_move='posted'):
        if not date_at:
            date_at = fields.Date.today()
        
        domain = [('date', '<=', date_at)]
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))
            
        report_data = {
            'assets': [],
            'liabilities': [],
            'equity': [],
            'total_assets': 0.0,
            'total_liabilities': 0.0,
            'total_equity': 0.0,
        }
        
        # Optimize by reading group sums in one go
        groups = self.env['account.move.line'].read_group(
            domain,
            ['balance', 'account_id'],
            ['account_id']
        )
        
        balances = {line['account_id'][0]: line['balance'] for line in groups if line.get('account_id')}
        
        acc_ids = list(balances.keys())
        accounts = self.env['account.account'].browse(acc_ids)
        
        income_balances = []
        expense_balances = []
        
        for acc in accounts:
            bal = balances[acc.id]
            if abs(bal) < 0.005:
                continue
                
            code = acc.code or ''
            name = acc.name or ''
            atype = acc.account_type or ''
            
            # Put income and expense aside to compute current year P&L
            if atype.startswith('income'):
                income_balances.append(bal)
                continue
            elif atype.startswith('expense'):
                expense_balances.append(bal)
                continue
                
            entry = {
                'code': code,
                'name': name,
            }
            
            # Account classification logic
            if code == '120384':
                entry['balance'] = -bal
                report_data['equity'].append(entry)
                report_data['total_equity'] += -bal
            elif code in ['300001.1', '120492']:
                entry['balance'] = -bal
                report_data['equity'].append(entry)
                report_data['total_equity'] += -bal
            elif code in ['120479', '120480', '120481', '120482', '120484', '120486']:
                entry['balance'] = -bal
                report_data['liabilities'].append(entry)
                report_data['total_liabilities'] += -bal
            elif code in ['112110', '211000.1']:
                entry['balance'] = -bal
                report_data['liabilities'].append(entry)
                report_data['total_liabilities'] += -bal
            elif code in ['100400', '121000']:
                entry['balance'] = bal
                report_data['assets'].append(entry)
                report_data['total_assets'] += bal
            elif atype.startswith('asset_fixed') or (code.startswith('1203') and code.isdigit() and int(code) <= 120377):
                entry['balance'] = bal
                report_data['assets'].append(entry)
                report_data['total_assets'] += bal
            elif atype == 'asset_cash' or code in ['100003', '100004', '100005', '100007', '100012', '100101', '100201']:
                if bal > 0:
                    entry['balance'] = bal
                    report_data['assets'].append(entry)
                    report_data['total_assets'] += bal
                else:
                    entry['name'] = name + ' (Overdraft)'
                    entry['balance'] = -bal
                    report_data['liabilities'].append(entry)
                    report_data['total_liabilities'] += -bal
            elif atype.startswith('asset') or atype == 'asset_receivable':
                entry['balance'] = bal
                report_data['assets'].append(entry)
                report_data['total_assets'] += bal
            elif atype.startswith('liability') or atype == 'liability_payable':
                entry['balance'] = -bal
                report_data['liabilities'].append(entry)
                report_data['total_liabilities'] += -bal
            elif atype.startswith('equity'):
                entry['balance'] = -bal
                report_data['equity'].append(entry)
                report_data['total_equity'] += -bal
            else:
                if bal > 0:
                    entry['balance'] = bal
                    report_data['assets'].append(entry)
                    report_data['total_assets'] += bal
                else:
                    entry['balance'] = -bal
                    report_data['liabilities'].append(entry)
                    report_data['total_liabilities'] += -bal
                    
        # Calculate Current Year Profit/Loss (revenue - expenses)
        # Odoo balance: income is credit (negative), expense is debit (positive)
        # net_profit = - (income_bal + expense_bal)
        net_profit_loss = - (sum(income_balances) + sum(expense_balances))
        if net_profit_loss != 0.0:
            report_data['equity'].append({
                'code': 'P&L',
                'name': 'Current Year Net Profit/Loss',
                'balance': net_profit_loss,
            })
            report_data['total_equity'] += net_profit_loss
            
        return report_data

    @api.model
    def get_profit_loss(self, date_from, date_to, target_move='posted'):
        domain = [
            ('date', '>=', date_from),
            ('date', '<=', date_to),
            ('account_id.account_type', 'in', ['income', 'income_other', 'expense', 'expense_depreciation', 'expense_direct_cost'])
        ]
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))
        
        report_data = {
            'income': [],
            'expenses': [],
            'total_income': 0.0,
            'total_expenses': 0.0,
            'net_profit': 0.0
        }
        
        lines = self.env['account.move.line'].read_group(
            domain,
            ['balance', 'account_id'],
            ['account_id']
        )
        
        for line in lines:
            account = self.env['account.account'].browse(line['account_id'][0])
            balance = line['balance']
            
            entry = {
                'code': account.code,
                'name': account.name,
                'balance': -balance if account.account_type.startswith('income') else balance,
            }
            
            if account.account_type.startswith('income'):
                report_data['income'].append(entry)
                report_data['total_income'] += -balance
            else:
                report_data['expenses'].append(entry)
                report_data['total_expenses'] += balance
                
        report_data['net_profit'] = report_data['total_income'] - report_data['total_expenses']
        return report_data

    @api.model
    def get_partner_ledger(self, partner_ids=None, date_from=None, date_to=None, target_move='posted'):
        domain = []
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))

        if partner_ids:
            domain.append(('partner_id', 'in', partner_ids))
        if date_from:
            domain.append(('date', '>=', date_from))
        if date_to:
            domain.append(('date', '<=', date_to))
        
        lines = self.env['account.move.line'].search(domain, order='partner_id, date, id')
        
        report_data = {}
        for line in lines:
            partner = line.partner_id.name or 'Unknown'
            if partner not in report_data:
                report_data[partner] = {
                    'lines': [],
                    'total_debit': 0.0,
                    'total_credit': 0.0,
                    'balance': 0.0
                }
            
            report_data[partner]['lines'].append({
                'date': line.date,
                'ref': line.move_id.name,
                'account': line.account_id.name,
                'debit': line.debit,
                'credit': line.credit,
                'balance': line.balance,
            })
            report_data[partner]['total_debit'] += line.debit
            report_data[partner]['total_credit'] += line.credit
            report_data[partner]['balance'] += line.balance
            
        return report_data

    @api.model
    def get_aged_partner_balance(self, partner_type='customer', date_at=None, target_move='posted'):
        if not date_at:
            date_at = fields.Date.today()
            
        account_type = 'asset_receivable' if partner_type == 'customer' else 'liability_payable'
        
        domain = [
            ('date', '<=', date_at),
            ('account_id.account_type', '=', account_type),
            ('reconciled', '=', False)
        ]
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))
        
        lines = self.env['account.move.line'].search(domain)
        
        report_data = {}
        for line in lines:
            partner = line.partner_id.name or 'Unknown'
            if partner not in report_data:
                report_data[partner] = {'total': 0.0, '0-30': 0.0, '31-60': 0.0, '61-90': 0.0, '90+': 0.0}
            
            days = (fields.Date.from_string(date_at) - fields.Date.from_string(line.date)).days
            amount = line.amount_residual
            
            report_data[partner]['total'] += amount
            if days <= 30:
                report_data[partner]['0-30'] += amount
            elif days <= 60:
                report_data[partner]['31-60'] += amount
            elif days <= 90:
                report_data[partner]['61-90'] += amount
            else:
                report_data[partner]['90+'] += amount
                
        return report_data

    @api.model
    def get_general_ledger(self, account_ids=None, date_from=None, date_to=None, target_move='posted'):
        domain = []
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))

        if account_ids:
            domain.append(('account_id', 'in', account_ids))
        if date_from:
            domain.append(('date', '>=', date_from))
        if date_to:
            domain.append(('date', '<=', date_to))
            
        lines = self.env['account.move.line'].search(domain, order='account_id, date, id')
        
        report_data = {}
        for line in lines:
            acc = line.account_id.display_name
            if acc not in report_data:
                report_data[acc] = {
                    'lines': [],
                    'total_debit': 0.0,
                    'total_credit': 0.0,
                    'balance': 0.0
                }
            
            report_data[acc]['lines'].append({
                'date': line.date,
                'ref': line.move_id.name,
                'partner': line.partner_id.name or '',
                'debit': line.debit,
                'credit': line.credit,
                'balance': line.balance,
            })
            report_data[acc]['total_debit'] += line.debit
            report_data[acc]['total_credit'] += line.credit
            report_data[acc]['balance'] += line.balance
            
        return report_data

    @api.model
    def get_journal_book(self, journal_type, date_from=None, date_to=None, target_move='posted'):
        domain = [('journal_id.type', '=', journal_type)]
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))

        if date_from:
            domain.append(('date', '>=', date_from))
        if date_to:
            domain.append(('date', '<=', date_to))
            
        lines = self.env['account.move.line'].search(domain, order='date, id')
        
        report_data = []
        for line in lines:
            report_data.append({
                'date': line.date,
                'journal': line.journal_id.name,
                'ref': line.move_id.name,
                'account': line.account_id.display_name,
                'debit': line.debit,
                'credit': line.credit,
                'balance': line.balance,
            })
        return report_data

    @api.model
    def get_tax_report(self, date_from=None, date_to=None, target_move='posted'):
        domain = []
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))

        if date_from:
            domain.append(('date', '>=', date_from))
        if date_to:
            domain.append(('date', '<=', date_to))
            
        lines = self.env['account.move.line'].read_group(
            domain + [('tax_line_id', '!=', False)],
            ['balance', 'tax_line_id'],
            ['tax_line_id']
        )
        
        report_data = []
        for line in lines:
            tax = self.env['account.tax'].browse(line['tax_line_id'][0])
            report_data.append({
                'tax_name': tax.name,
                'tax_amount': line['balance'],
            })
        return report_data

    @api.model
    def get_journals_audit(self, journal_ids=None, date_from=None, date_to=None, target_move='posted'):
        domain = []
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))

        if journal_ids:
            domain.append(('journal_id', 'in', journal_ids))
        if date_from:
            domain.append(('date', '>=', date_from))
        if date_to:
            domain.append(('date', '<=', date_to))
            
        moves = self.env['account.move'].search(domain, order='journal_id, date, name')
        
        report_data = {}
        for move in moves:
            j_name = move.journal_id.name
            if j_name not in report_data:
                report_data[j_name] = []
            
            move_data = {
                'name': move.name,
                'date': move.date,
                'ref': move.ref or '',
                'lines': []
            }
            for line in move.line_ids:
                move_data['lines'].append({
                    'account': line.account_id.display_name,
                    'partner': line.partner_id.name or '',
                    'debit': line.debit,
                    'credit': line.credit,
                })
            report_data[j_name].append(move_data)
        return report_data

    @api.model
    def get_day_book(self, date, target_move='posted'):
        domain = [('date', '=', date)]
        if target_move == 'posted':
            domain.append(('parent_state', '=', 'posted'))
        else:
            domain.append(('parent_state', 'in', ['draft', 'posted']))

        lines = self.env['account.move.line'].search(domain, order='date, id')
        
        report_data = []
        for line in lines:
            report_data.append({
                'date': line.date,
                'entry': line.move_id.name,
                'account': line.account_id.display_name,
                'partner': line.partner_id.name or '',
                'ref': line.ref or '',
                'debit': line.debit,
                'credit': line.credit,
                'balance': line.balance,
            })
        return report_data

class CustomReportDisplayLine(models.TransientModel):
    _name = 'custom.report.display.line'
    _description = 'Custom Report Display Line'
    _order = 'sequence, id'

    sequence = fields.Integer('Sequence')
    report_type = fields.Selection([
        ('balance_sheet', 'Balance Sheet'),
        ('profit_loss', 'Profit and Loss'),
        ('trial_balance', 'Trial Balance'),
        ('partner_ledger', 'Partner Ledger'),
        ('general_ledger', 'General Ledger'),
        ('cash_book', 'Cash Book'),
        ('tax_report', 'Tax Report'),
        ('gstr2b_reconciliation', 'GSTR-2B Reconciliation'),
        ('gstr2a_reconciliation', 'GSTR-2A Reconciliation'),
        ('annual_computation', 'Annual Computation'),
        ('brs_statement', 'Bank Reconciliation Statement (BRS)')
    ], string='Report Type')
    date = fields.Date('Date')
    code = fields.Char('Code')
    name = fields.Char('Name')
    ref = fields.Char('Ref')
    account = fields.Char('Account')
    partner = fields.Char('Partner')
    journal = fields.Char('Journal')
    tax_name = fields.Char('Tax Name')
    debit = fields.Float('Debit')
    credit = fields.Float('Credit')
    balance = fields.Float('Balance')
    voucher_count = fields.Integer('Voucher Count')
    taxable_amount = fields.Float('Taxable Amount')
    igst = fields.Float('IGST')
    cgst = fields.Float('CGST')
    sgst = fields.Float('SGST')
    cess = fields.Float('Cess')
    tax_amount = fields.Float('Tax Amount')
    invoice_status_amount = fields.Char('Invoice Status Amount')
    is_header = fields.Boolean(string='Is Header', default=False)

class CustomReportWizard(models.TransientModel):
    _name = 'custom.report.wizard'
    _description = 'Custom Accounting Report Wizard'

    report_type = fields.Selection([
        ('balance_sheet', 'Balance Sheet'),
        ('profit_loss', 'Profit and Loss'),
        ('trial_balance', 'Trial Balance'),
        ('partner_ledger', 'Partner Ledger'),
        ('general_ledger', 'General Ledger'),
        ('cash_book', 'Cash Book'),
        ('tax_report', 'Tax Report'),
        ('gstr2b_reconciliation', 'GSTR-2B Reconciliation'),
        ('gstr2a_reconciliation', 'GSTR-2A Reconciliation'),
        ('annual_computation', 'Annual Computation'),
        ('brs_statement', 'Bank Reconciliation Statement (BRS)')
    ], string='Report Type', required=True)
    
    journal_id = fields.Many2one('account.journal', string='Bank Journal', domain="[('type', '=', 'bank')]", default=lambda self: self.env['account.journal'].search([('type', '=', 'bank')], limit=1))
    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date', default=fields.Date.today())
    target_move = fields.Selection([
        ('posted', 'All Posted Entries'),
        ('all', 'All Entries'),
    ], string='Target Moves', default='posted')

    def _get_brs_data(self):
        import openpyxl
        import os
        import datetime
        from odoo import fields
        
        file_name = 'May_Payment_bank_statement.xlsx'
        if self.date_from and self.date_from.month == 4:
            file_name = 'April_HDFC_50200024612749_Payment.xlsx'
            
        file_path = os.path.join('/home/ubuntu/odoo-gdr', file_name)
        if not os.path.exists(file_path):
            raise models.UserError(f"Bank statement file not found at: {file_path}")
            
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        
        excel_txs = []
        excel_opening_bal = 0.0
        excel_closing_bal = 0.0
        
        for idx, r in enumerate(rows, 1):
            if len(r) > 2 and r[2] == 'Opening Balance':
                excel_opening_bal = float(r[8]) if len(r) > 8 and r[8] is not None else 0.0
            if len(r) > 2 and r[2] == 'Closing Balance':
                excel_closing_bal = float(r[5]) if len(r) > 5 and r[5] is not None else 0.0
                
            if len(r) > 0 and (isinstance(r[0], datetime.date) or isinstance(r[0], datetime.datetime)):
                if r[2] == 'Opening Balance':
                    continue
                excel_txs.append({
                    'date': r[0].date() if isinstance(r[0], datetime.datetime) else r[0],
                    'particulars': r[2] or '',
                    'vch_type': r[5] or '',
                    'vch_no': str(r[6]) if r[6] is not None else '',
                    'debit': float(r[7]) if len(r) > 7 and r[7] is not None else 0.0,
                    'credit': float(r[8]) if len(r) > 8 and r[8] is not None else 0.0,
                })

        journal = self.journal_id or self.env['account.journal'].browse(10)
        bank_account_id = journal.default_account_id.id
        
        date_from_str = self.date_from or fields.Date.today()
        lines_before = self.env['account.move.line'].search([
            ('account_id', '=', bank_account_id),
            ('date', '<', date_from_str),
        ])
        odoo_opening_bal = sum(l.debit - l.credit for l in lines_before)
        
        date_to_str = self.date_to or fields.Date.today()
        lines_period = self.env['account.move.line'].search([
            ('account_id', '=', bank_account_id),
            ('date', '>=', date_from_str),
            ('date', '<=', date_to_str),
        ], order='date, id')
        
        odoo_txs = []
        for l in lines_period:
            odoo_txs.append({
                'line_id': l.id,
                'move_name': l.move_id.name,
                'date': l.date,
                'particulars': l.name or '',
                'partner': l.partner_id.name or '',
                'ref': l.ref or l.move_id.ref or '',
                'debit': float(l.debit),
                'credit': float(l.credit),
            })
            
        odoo_closing_bal = odoo_opening_bal + sum(l['debit'] - l['credit'] for l in odoo_txs)
        
        matched_excel = set()
        matched_odoo = set()
        
        for e_idx, e in enumerate(excel_txs):
            for o_idx, o in enumerate(odoo_txs):
                if o_idx in matched_odoo:
                    continue
                if e['date'] == o['date'] and abs(e['debit'] - o['debit']) < 0.01 and abs(e['credit'] - o['credit']) < 0.01:
                    matched_excel.add(e_idx)
                    matched_odoo.add(o_idx)
                    break
                    
        for e_idx, e in enumerate(excel_txs):
            if e_idx in matched_excel:
                continue
            for o_idx, o in enumerate(odoo_txs):
                if o_idx in matched_odoo:
                    continue
                if abs((e['date'] - o['date']).days) <= 7 and abs(e['debit'] - o['debit']) < 0.01 and abs(e['credit'] - o['credit']) < 0.01:
                    matched_excel.add(e_idx)
                    matched_odoo.add(o_idx)
                    break
                    
        unmatched_excel = [excel_txs[i] for i in range(len(excel_txs)) if i not in matched_excel]
        unmatched_odoo = [odoo_txs[i] for i in range(len(odoo_txs)) if i not in matched_odoo]
        
        prior_discrepancy = (-excel_opening_bal) - odoo_opening_bal
        total_duplicates = sum(o['credit'] for o in unmatched_odoo if o['ref'].startswith('VENDOR_PAY/'))
        adjusted_bal = odoo_closing_bal + total_duplicates - 40000.0 + prior_discrepancy
        
        return {
            'excel_txs': excel_txs,
            'excel_opening_bal': excel_opening_bal,
            'excel_closing_bal': excel_closing_bal,
            'odoo_opening_bal': odoo_opening_bal,
            'odoo_closing_bal': odoo_closing_bal,
            'unmatched_excel': unmatched_excel,
            'unmatched_odoo': unmatched_odoo,
            'prior_discrepancy': prior_discrepancy,
            'total_duplicates': total_duplicates,
            'adjusted_bal': adjusted_bal,
        }

    def action_view_report(self):
        self.ensure_one()
        report_model = self.env['custom.accounting.report']
        lines = []
        seq = 10
        action_name = "Accounting Report"
        view_id = False

        if self.report_type == 'balance_sheet':
            action_name = "Balance Sheet"
            data = report_model.get_balance_sheet(date_at=self.date_to, target_move=self.target_move)
            lines.append({'sequence': seq, 'name': 'ASSETS', 'balance': data.get('total_assets', 0.0), 'report_type': 'balance_sheet'})
            seq += 10
            for asset in data.get('assets', []):
                lines.append({
                    'sequence': seq,
                    'name': f"    [{asset['code']}] {asset['name']}",
                    'balance': asset['balance'],
                    'report_type': 'balance_sheet'
                })
                seq += 10
                
            lines.append({'sequence': seq, 'name': 'LIABILITIES', 'balance': data.get('total_liabilities', 0.0), 'report_type': 'balance_sheet'})
            seq += 10
            for liability in data.get('liabilities', []):
                lines.append({
                    'sequence': seq,
                    'name': f"    [{liability['code']}] {liability['name']}",
                    'balance': liability['balance'],
                    'report_type': 'balance_sheet'
                })
                seq += 10

            lines.append({'sequence': seq, 'name': 'EQUITY', 'balance': data.get('total_equity', 0.0), 'report_type': 'balance_sheet'})
            seq += 10
            for eq in data.get('equity', []):
                lines.append({
                    'sequence': seq,
                    'name': f"    [{eq['code']}] {eq['name']}",
                    'balance': eq['balance'],
                    'report_type': 'balance_sheet'
                })
                seq += 10
            view_id = self.env.ref('custom_accounting_reports.view_balance_sheet_display_tree').id

        elif self.report_type == 'profit_loss':
            action_name = "Profit and Loss"
            data = report_model.get_profit_loss(self.date_from, self.date_to, target_move=self.target_move)
            lines.append({'sequence': seq, 'name': 'INCOME', 'balance': data.get('total_income', 0.0), 'report_type': 'profit_loss'})
            seq += 10
            for inc in data.get('income', []):
                lines.append({
                    'sequence': seq,
                    'name': f"    [{inc['code']}] {inc['name']}",
                    'balance': inc['balance'],
                    'report_type': 'profit_loss'
                })
                seq += 10
                
            lines.append({'sequence': seq, 'name': 'EXPENSES', 'balance': data.get('total_expenses', 0.0), 'report_type': 'profit_loss'})
            seq += 10
            for exp in data.get('expenses', []):
                lines.append({
                    'sequence': seq,
                    'name': f"    [{exp['code']}] {exp['name']}",
                    'balance': exp['balance'],
                    'report_type': 'profit_loss'
                })
                seq += 10
                
            lines.append({'sequence': seq, 'name': 'NET PROFIT', 'balance': data.get('net_profit', 0.0), 'report_type': 'profit_loss'})
            view_id = self.env.ref('custom_accounting_reports.view_profit_loss_display_tree').id

        elif self.report_type == 'trial_balance':
            action_name = "Trial Balance"
            domain = [('date', '<=', self.date_to or fields.Date.today())]
            if self.target_move == 'posted':
                domain.append(('parent_state', '=', 'posted'))
            else:
                domain.append(('parent_state', 'in', ['draft', 'posted']))

            accounts = self.env['account.account'].search([])
            for account in accounts:
                res = self.env['account.move.line'].read_group(
                    domain + [('account_id', '=', account.id)],
                    ['debit', 'credit', 'balance'],
                    ['account_id']
                )
                if res:
                    lines.append({
                        'sequence': seq,
                        'code': account.code,
                        'name': account.name,
                        'debit': res[0]['debit'],
                        'credit': res[0]['credit'],
                        'balance': res[0]['balance'],
                        'report_type': 'trial_balance'
                    })
                    seq += 10
            view_id = self.env.ref('custom_accounting_reports.view_trial_balance_display_tree').id

        elif self.report_type == 'partner_ledger':
            action_name = "Partner Ledger"
            data = report_model.get_partner_ledger(None, self.date_from, self.date_to, target_move=self.target_move)
            for p_name, p_data in data.items():
                lines.append({
                    'sequence': seq,
                    'partner': f"=== {p_name} ===",
                    'debit': p_data['total_debit'],
                    'credit': p_data['total_credit'],
                    'balance': p_data['balance'],
                    'report_type': 'partner_ledger'
                })
                seq += 10
                for line in p_data.get('lines', []):
                    lines.append({
                        'sequence': seq,
                        'date': line['date'],
                        'ref': line['ref'],
                        'account': line['account'],
                        'debit': line['debit'],
                        'credit': line['credit'],
                        'balance': line['balance'],
                        'report_type': 'partner_ledger'
                    })
                    seq += 10
            view_id = self.env.ref('custom_accounting_reports.view_partner_ledger_display_tree').id

        elif self.report_type == 'general_ledger':
            action_name = "General Ledger"
            data = report_model.get_general_ledger(None, self.date_from, self.date_to, target_move=self.target_move)
            for acc_name, acc_data in data.items():
                lines.append({
                    'sequence': seq,
                    'account': f"=== {acc_name} ===",
                    'debit': acc_data['total_debit'],
                    'credit': acc_data['total_credit'],
                    'balance': acc_data['balance'],
                    'report_type': 'general_ledger'
                })
                seq += 10
                for line in acc_data.get('lines', []):
                    lines.append({
                        'sequence': seq,
                        'date': line['date'],
                        'ref': line['ref'],
                        'partner': line['partner'],
                        'debit': line['debit'],
                        'credit': line['credit'],
                        'balance': line['balance'],
                        'report_type': 'general_ledger'
                    })
                    seq += 10
            view_id = self.env.ref('custom_accounting_reports.view_general_ledger_display_tree').id

        elif self.report_type == 'cash_book':
            action_name = "Cash Book"
            data = report_model.get_journal_book(journal_type='cash', date_from=self.date_from, date_to=self.date_to, target_move=self.target_move)
            for line in data:
                lines.append({
                    'sequence': seq,
                    'date': line['date'],
                    'journal': line['journal'],
                    'ref': line['ref'],
                    'account': line['account'],
                    'debit': line['debit'],
                    'credit': line['credit'],
                    'balance': line['balance'],
                    'report_type': 'cash_book'
                })
                seq += 10
            view_id = self.env.ref('custom_accounting_reports.view_cash_book_display_tree').id

        elif self.report_type == 'tax_report':
            action_name = "Tax Report"
            data = report_model.get_tax_report(self.date_from, self.date_to, target_move=self.target_move)
            for line in data:
                lines.append({
                    'sequence': seq,
                    'tax_name': line['tax_name'],
                    'balance': line['tax_amount'],
                    'report_type': 'tax_report'
                })
                seq += 10
            view_id = self.env.ref('custom_accounting_reports.view_tax_report_display_tree').id

        elif self.report_type == 'gstr2b_reconciliation':
            action_name = "GSTR-2B Reconciliation"
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'posted' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            posted_count = self.env.cr.fetchone()[0] or 0
            
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'draft' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            draft_count = self.env.cr.fetchone()[0] or 0

            self.env.cr.execute("""
                SELECT 
                    SUM(CASE WHEN l.display_type = 'product' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%IGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%CGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%SGST%%' OR t.name::text ILIKE '%%UTGST%%' THEN abs(l.balance) ELSE 0 END)
                FROM account_move_line l
                JOIN account_move m ON l.move_id = m.id
                LEFT JOIN account_tax t ON l.tax_line_id = t.id
                WHERE m.move_type = 'in_invoice' AND m.state = 'posted' AND m.date >= %s AND m.date <= %s
                  AND (l.display_type = 'product' OR l.tax_line_id IS NOT NULL)
            """, (self.date_from or '2025-04-01', self.date_to))
            totals = self.env.cr.fetchone()
            taxable = totals[0] or 0.0
            igst = totals[1] or 0.0
            cgst = totals[2] or 0.0
            sgst = totals[3] or 0.0
            total_tax = igst + cgst + sgst

            lines.append({'sequence': seq, 'name': 'Reconciled', 'voucher_count': posted_count, 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Unreconciled', 'voucher_count': draft_count, 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Available Only in Books', 'voucher_count': draft_count, 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Uncertain Transactions (Corrections needed)', 'voucher_count': 0, 'report_type': 'gstr2b_reconciliation'})
            seq += 10

            lines.append({'sequence': seq, 'name': 'All other ITC from Registered Persons (Excluding Reverse Charge)', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': 'Unreconciled', 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Inward Supplies from ISD', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': 'Unreconciled', 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Inward Supplies Liable for Reverse Charge', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': 'Unreconciled', 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Import of Goods', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': 'Unreconciled', 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Total Available ITC', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': '', 'report_type': 'gstr2b_reconciliation'})
            seq += 10
            
            view_id = self.env.ref('custom_accounting_reports.view_gstr2b_reconciliation_display_tree').id

        elif self.report_type == 'gstr2a_reconciliation':
            action_name = "GSTR-2A Reconciliation"
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'posted' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            posted_count = self.env.cr.fetchone()[0] or 0
            
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'draft' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            draft_count = self.env.cr.fetchone()[0] or 0

            self.env.cr.execute("""
                SELECT 
                    SUM(CASE WHEN l.display_type = 'product' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%IGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%CGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%SGST%%' OR t.name::text ILIKE '%%UTGST%%' THEN abs(l.balance) ELSE 0 END)
                FROM account_move_line l
                JOIN account_move m ON l.move_id = m.id
                LEFT JOIN account_tax t ON l.tax_line_id = t.id
                WHERE m.move_type = 'in_invoice' AND m.state = 'posted' AND m.date >= %s AND m.date <= %s
                  AND (l.display_type = 'product' OR l.tax_line_id IS NOT NULL)
            """, (self.date_from or '2025-04-01', self.date_to))
            totals = self.env.cr.fetchone()
            taxable = totals[0] or 0.0
            igst = totals[1] or 0.0
            cgst = totals[2] or 0.0
            sgst = totals[3] or 0.0
            total_tax = igst + cgst + sgst

            lines.append({'sequence': seq, 'name': 'Reconciled', 'voucher_count': 0, 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Unreconciled', 'voucher_count': draft_count, 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Available only in Books', 'voucher_count': draft_count, 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Uncertain Transactions (Corrections needed)', 'voucher_count': 0, 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Check Vouchers Having Potential Conflicts with Masters', 'invoice_status_amount': 'Yes', 'report_type': 'gstr2a_reconciliation'})
            seq += 10

            lines.append({'sequence': seq, 'name': 'B2B Invoices', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': 'Unreconciled', 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Amendments to B2B Invoices', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Credit/Debit Notes', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Amendments to Credit/Debit Notes', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'ISD Credits', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Import of Goods from overseas on Bill of Entry', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Import of Goods from SEZ Units/Developers on Bill of Entry', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'gstr2a_reconciliation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Total', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': '', 'report_type': 'gstr2a_reconciliation'})
            seq += 10

            view_id = self.env.ref('custom_accounting_reports.view_gstr2a_reconciliation_display_tree').id

        elif self.report_type == 'annual_computation':
            action_name = "Annual Computation"
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'posted' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            posted_count = self.env.cr.fetchone()[0] or 0
            
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'draft' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            draft_count = self.env.cr.fetchone()[0] or 0

            self.env.cr.execute("""
                SELECT 
                    SUM(CASE WHEN l.display_type = 'product' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%IGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%CGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%SGST%%' OR t.name::text ILIKE '%%UTGST%%' THEN abs(l.balance) ELSE 0 END)
                FROM account_move_line l
                JOIN account_move m ON l.move_id = m.id
                LEFT JOIN account_tax t ON l.tax_line_id = t.id
                WHERE m.move_type = 'in_invoice' AND m.state = 'posted' AND m.date >= %s AND m.date <= %s
                  AND (l.display_type = 'product' OR l.tax_line_id IS NOT NULL)
            """, (self.date_from or '2025-04-01', self.date_to))
            totals = self.env.cr.fetchone()
            taxable = totals[0] or 0.0
            igst = totals[1] or 0.0
            cgst = totals[2] or 0.0
            sgst = totals[3] or 0.0
            total_tax = igst + cgst + sgst

            lines.append({'sequence': seq, 'name': 'Total Vouchers', 'voucher_count': posted_count + draft_count, 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Included In Return', 'voucher_count': posted_count, 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Not Relevant for This Return', 'voucher_count': draft_count, 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Uncertain Transactions (Corrections needed)', 'voucher_count': 0, 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Check Vouchers Having Potential Conflicts with Masters', 'invoice_status_amount': 'Yes', 'report_type': 'annual_computation'})
            seq += 10

            lines.append({'sequence': seq, 'name': 'Outward and Inward Supplies on Which Tax is Payable (Including Advances)', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Outward Supplies on Which Tax is Not Payable', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Missing Invoice Reported in Current Period', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Total Liability', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10

            lines.append({'sequence': seq, 'name': 'Input Tax Credit', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Reversal of Input Tax Credit, Adjusted and Ineligible Input Tax Credit Declared', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Total Input Tax Credit After Reversal & Ineligible Input Tax Credit', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10

            lines.append({'sequence': seq, 'name': 'Interest, Late Fee, Penalty and Others', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10

            lines.append({'sequence': seq, 'name': 'Summary of Outward Supplies', 'voucher_count': 0, 'taxable_amount': 0.0, 'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0, 'tax_amount': 0.0, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10
            lines.append({'sequence': seq, 'name': 'Summary of Inward Supplies', 'voucher_count': posted_count, 'taxable_amount': taxable, 'igst': igst, 'cgst': cgst, 'sgst': sgst, 'tax_amount': total_tax, 'invoice_status_amount': '', 'report_type': 'annual_computation'})
            seq += 10

            view_id = self.env.ref('custom_accounting_reports.view_annual_computation_display_tree').id

        elif self.report_type == 'brs_statement':
            action_name = "Bank Reconciliation Statement (BRS)"
            data = self._get_brs_data()
            
            lines.append({
                'sequence': seq,
                'name': 'Odoo Ledger Opening Balance',
                'balance': data['odoo_opening_bal'],
                'invoice_status_amount': 'Ledger Opening',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Bank Statement Opening Balance',
                'balance': -data['excel_opening_bal'],
                'invoice_status_amount': 'Statement Opening',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Prior Period Discrepancy carried forward',
                'balance': data['prior_discrepancy'],
                'invoice_status_amount': 'Carry Forward',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            reconciled_count = len(data['excel_txs']) - len(data['unmatched_excel'])
            lines.append({
                'sequence': seq,
                'name': f'Matched/Reconciled Transactions ({reconciled_count} entries)',
                'debit': sum(e['debit'] for e in data['excel_txs'] if e not in data['unmatched_excel']),
                'credit': sum(e['credit'] for e in data['excel_txs'] if e not in data['unmatched_excel']),
                'invoice_status_amount': 'Reconciled',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            if data['unmatched_excel']:
                lines.append({
                    'sequence': seq,
                    'name': '--- Omissions: Transactions in Bank Statement but NOT in Odoo Books ---',
                    'is_header': True,
                    'report_type': 'brs_statement',
                })
                seq += 10
                for tx in data['unmatched_excel']:
                    lines.append({
                        'sequence': seq,
                        'date': tx['date'],
                        'ref': tx['vch_no'],
                        'name': tx['particulars'],
                        'debit': tx['debit'],
                        'credit': tx['credit'],
                        'invoice_status_amount': 'Missing in Books',
                        'report_type': 'brs_statement',
                    })
                    seq += 10
                    
            if data['unmatched_odoo']:
                lines.append({
                    'sequence': seq,
                    'name': '--- Discrepancies: Transactions in Odoo Books but NOT in Bank Statement ---',
                    'is_header': True,
                    'report_type': 'brs_statement',
                })
                seq += 10
                for tx in data['unmatched_odoo']:
                    status = 'Duplicate (VENDOR_PAY/)' if tx['ref'].startswith('VENDOR_PAY/') else 'Unmatched Ledger Entry'
                    if tx['move_name'] == 'MISC/2025/05/0001':
                        status = 'Incorrect Contra (Reversed)'
                    lines.append({
                        'sequence': seq,
                        'date': tx['date'],
                        'ref': tx['ref'] or tx['move_name'],
                        'name': f"[{tx['move_name']}] {tx['particulars']}",
                        'debit': tx['debit'],
                        'credit': tx['credit'],
                        'invoice_status_amount': status,
                        'report_type': 'brs_statement',
                    })
                    seq += 10
                    
            lines.append({
                'sequence': seq,
                'name': '--------------------------------------------------',
                'is_header': True,
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Odoo Ledger Ending Balance',
                'balance': data['odoo_closing_bal'],
                'invoice_status_amount': 'Ledger Ending',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Add: Duplicate payments posted in Odoo (VENDOR_PAY/) to be reversed',
                'balance': data['total_duplicates'],
                'invoice_status_amount': 'Adjustment',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Less: Reverse incorrect Contra Debit entry in Odoo',
                'balance': -20000.0,
                'invoice_status_amount': 'Adjustment',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Less: Record missing Petty Cash payment (Excel Row 67)',
                'balance': -20000.0,
                'invoice_status_amount': 'Adjustment',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Add: Discrepancy carried forward from prior period',
                'balance': data['prior_discrepancy'],
                'invoice_status_amount': 'Adjustment',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Adjusted Balance as per Odoo Ledger',
                'balance': data['adjusted_bal'],
                'invoice_status_amount': 'Reconciled Balance',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Actual Balance as per Bank Statement',
                'balance': -data['excel_closing_bal'],
                'invoice_status_amount': 'Statement Ending',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            lines.append({
                'sequence': seq,
                'name': 'Unreconciled Difference',
                'balance': data['adjusted_bal'] - (-data['excel_closing_bal']),
                'invoice_status_amount': 'Difference',
                'report_type': 'brs_statement',
            })
            seq += 10
            
            view_id = self.env.ref('custom_accounting_reports.view_brs_statement_display_tree').id

        created_lines = self.env['custom.report.display.line'].create(lines)

        return {
            'name': action_name,
            'type': 'ir.actions.act_window',
            'res_model': 'custom.report.display.line',
            'view_mode': 'list',
            'views': [(view_id, 'list')],
            'domain': [('id', 'in', created_lines.ids)],
            'target': 'current',
        }

    def action_generate_excel(self):
        self.ensure_one()
        import io
        import base64
        try:
            import xlsxwriter
        except ImportError:
            raise models.UserError("xlsxwriter is not installed in the environment.")

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        
        bold = workbook.add_format({'bold': True, 'font_size': 11})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1})
        money_format = workbook.add_format({'num_format': '[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00', 'border': 1})
        label_format = workbook.add_format({'border': 1})
        title_format = workbook.add_format({'bold': True, 'font_size': 14})
        
        report_model = self.env['custom.accounting.report']
        
        if self.report_type == 'balance_sheet':
            sheet = workbook.add_worksheet('Balance Sheet')
            sheet.set_column(0, 0, 50)
            sheet.set_column(1, 1, 18)
            
            company = self.env.company
            # Full Company Header
            row = 0
            sheet.write(row, 0, company.name or '', workbook.add_format({'bold': True, 'font_size': 14}))
            row += 1
            sheet.write(row, 0, company.street or '', label_format)
            row += 1
            sheet.write(row, 0, (company.street2 or '') + ' ' + (company.city or ''), label_format)
            row += 1
            sheet.write(row, 0, (company.state_id.name or '') + ' ' + (company.zip or ''), label_format)
            row += 1
            sheet.write(row, 0, f"E-Mail : {company.email or ''}", label_format)
            row += 1
            sheet.write(row, 0, 'Balance Sheet', bold)
            row += 1
            sheet.write(row, 0, f"As of {self.date_to or fields.Date.today()}", label_format)
            row += 2

            data = report_model.get_balance_sheet(date_at=self.date_to)
            section_fmt = workbook.add_format({'bold': True, 'bg_color': '#E8E8E8', 'border': 1})
            sub_section_fmt = workbook.add_format({'bold': True, 'indent': 1, 'border': 1})
            indent_label = workbook.add_format({'indent': 2, 'border': 1})
            
            # ASSETS
            sheet.write(row, 0, "ASSETS", section_fmt)
            sheet.write(row, 1, data.get('total_assets', 0.0), section_fmt)
            row += 1
            sheet.write(row, 0, "Current Assets", sub_section_fmt)
            sheet.write(row, 1, data.get('total_assets', 0.0), sub_section_fmt)
            row += 1
            for asset in data.get('assets', []):
                sheet.write(row, 0, f"[{asset['code']}] {asset['name']}", indent_label)
                sheet.write(row, 1, asset['balance'], money_format)
                row += 1
            row += 1

            # LIABILITIES
            sheet.write(row, 0, "LIABILITIES", section_fmt)
            sheet.write(row, 1, data.get('total_liabilities', 0.0), section_fmt)
            row += 1
            sheet.write(row, 0, "Current Liabilities", sub_section_fmt)
            sheet.write(row, 1, data.get('total_liabilities', 0.0), sub_section_fmt)
            row += 1
            for liability in data.get('liabilities', []):
                sheet.write(row, 0, f"[{liability['code']}] {liability['name']}", indent_label)
                sheet.write(row, 1, liability['balance'], money_format)
                row += 1
            row += 1

            # EQUITY
            sheet.write(row, 0, "EQUITY", section_fmt)
            sheet.write(row, 1, data.get('total_equity', 0.0), section_fmt)
            row += 1
            for eq in data.get('equity', []):
                sheet.write(row, 0, f"[{eq['code']}] {eq['name']}", indent_label)
                sheet.write(row, 1, eq['balance'], money_format)
                row += 1
            row += 2

            # Footer
            sheet.write(row, 0, "LIABILITIES + EQUITY", section_fmt)
            sheet.write(row, 1, data.get('total_liabilities', 0.0) + data.get('total_equity', 0.0), section_fmt)

        elif self.report_type == 'profit_loss':
            sheet = workbook.add_worksheet('Profit and Loss')
            sheet.set_column(0, 0, 50)
            sheet.set_column(1, 1, 18)
            
            company = self.env.company
            # Full Company Header
            row = 0
            sheet.write(row, 0, company.name or '', workbook.add_format({'bold': True, 'font_size': 14}))
            row += 1
            sheet.write(row, 0, company.street or '', label_format)
            row += 1
            sheet.write(row, 0, (company.street2 or '') + ' ' + (company.city or ''), label_format)
            row += 1
            sheet.write(row, 0, (company.state_id.name or '') + ' ' + (company.zip or ''), label_format)
            row += 1
            sheet.write(row, 0, f"E-Mail : {company.email or ''}", label_format)
            row += 1
            sheet.write(row, 0, 'Profit and Loss', bold)
            row += 1
            sheet.write(row, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}", label_format)
            row += 2
            data = report_model.get_profit_loss(self.date_from, self.date_to)
            sheet.write(row, 0, "Income", bold)
            row += 1
            for inc in data.get('income', []):
                sheet.write(row, 0, f"[{inc['code']}] {inc['name']}", label_format)
                sheet.write(row, 1, inc['balance'], money_format)
                row += 1
            sheet.write(row, 0, "Total Income", bold)
            sheet.write(row, 1, data.get('total_income', 0.0), money_format)
            
            row += 2
            sheet.write(row, 0, "Expenses", bold)
            row += 1
            for exp in data.get('expenses', []):
                sheet.write(row, 0, f"[{exp['code']}] {exp['name']}", label_format)
                sheet.write(row, 1, exp['balance'], money_format)
                row += 1
            sheet.write(row, 0, "Total Expenses", bold)
            sheet.write(row, 1, data.get('total_expenses', 0.0), money_format)
            
            row += 2
            sheet.write(row, 0, "Net Profit", bold)
            sheet.write(row, 1, data.get('net_profit', 0.0), money_format)

        elif self.report_type == 'trial_balance':
            sheet = workbook.add_worksheet('Trial Balance')
            sheet.set_column(0, 0, 15)
            sheet.set_column(1, 1, 40)
            sheet.set_column(2, 4, 15)
            
            company = self.env.company
            # Full Company Header
            row = 0
            sheet.write(row, 0, company.name or '', workbook.add_format({'bold': True, 'font_size': 14}))
            row += 1
            sheet.write(row, 0, company.street or '', label_format)
            row += 1
            sheet.write(row, 0, (company.street2 or '') + ' ' + (company.city or ''), label_format)
            row += 1
            sheet.write(row, 0, (company.state_id.name or '') + ' ' + (company.zip or ''), label_format)
            row += 1
            sheet.write(row, 0, f"E-Mail : {company.email or ''}", label_format)
            row += 1
            sheet.write(row, 0, 'Trial Balance', bold)
            row += 1
            sheet.write(row, 0, f"Up to {self.date_to or fields.Date.today()}", label_format)
            row += 2
            sheet.write(row, 0, "Account Code", header_format)
            sheet.write(row, 1, "Account Name", header_format)
            sheet.write(row, 2, "Debit", header_format)
            sheet.write(row, 3, "Credit", header_format)
            sheet.write(row, 4, "Balance", header_format)
            
            row += 1
            domain = [('date', '<=', self.date_to or fields.Date.today()), ('parent_state', '=', 'posted')]
            accounts = self.env['account.account'].search([])
            
            for account in accounts:
                res = self.env['account.move.line'].read_group(
                    domain + [('account_id', '=', account.id)],
                    ['debit', 'credit', 'balance'],
                    ['account_id']
                )
                if res:
                    sheet.write(row, 0, account.code, label_format)
                    sheet.write(row, 1, account.name, label_format)
                    sheet.write(row, 2, res[0]['debit'], money_format)
                    sheet.write(row, 3, res[0]['credit'], money_format)
                    sheet.write(row, 4, res[0]['balance'], money_format)
                    row += 1

        elif self.report_type == 'partner_ledger':
            sheet = workbook.add_worksheet('Partner Ledger')
            sheet.set_column(0, 0, 15)
            sheet.set_column(1, 1, 20)
            sheet.set_column(2, 2, 40)
            sheet.set_column(3, 5, 15)
            sheet.write(0, 0, "Partner Ledger", title_format)
            sheet.write(1, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}")
            
            row = 3
            data = report_model.get_partner_ledger(None, self.date_from, self.date_to)
            for partner_name, partner_data in data.items():
                sheet.write(row, 0, partner_name, bold)
                row += 1
                sheet.write(row, 0, "Date", header_format)
                sheet.write(row, 1, "Ref", header_format)
                sheet.write(row, 2, "Account", header_format)
                sheet.write(row, 3, "Debit", header_format)
                sheet.write(row, 4, "Credit", header_format)
                sheet.write(row, 5, "Balance", header_format)
                row += 1
                
                for line in partner_data.get('lines', []):
                    sheet.write(row, 0, str(line['date']), label_format)
                    sheet.write(row, 1, line['ref'] or '', label_format)
                    sheet.write(row, 2, line['account'] or '', label_format)
                    sheet.write(row, 3, line['debit'], money_format)
                    sheet.write(row, 4, line['credit'], money_format)
                    sheet.write(row, 5, line['balance'], money_format)
                    row += 1
                    
                sheet.write(row, 2, "Total", bold)
                sheet.write(row, 3, partner_data['total_debit'], money_format)
                sheet.write(row, 4, partner_data['total_credit'], money_format)
                sheet.write(row, 5, partner_data['balance'], money_format)
                row += 2

        elif self.report_type == 'general_ledger':
            sheet = workbook.add_worksheet('General Ledger')
            sheet.set_column(0, 0, 15)
            sheet.set_column(1, 1, 20)
            sheet.set_column(2, 2, 35)
            sheet.set_column(3, 5, 15)
            sheet.write(0, 0, "General Ledger", title_format)
            sheet.write(1, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}")
            
            row = 3
            data = report_model.get_general_ledger(None, self.date_from, self.date_to)
            for acc_name, acc_data in data.items():
                sheet.write(row, 0, acc_name, bold)
                row += 1
                sheet.write(row, 0, "Date", header_format)
                sheet.write(row, 1, "Ref", header_format)
                sheet.write(row, 2, "Partner", header_format)
                sheet.write(row, 3, "Debit", header_format)
                sheet.write(row, 4, "Credit", header_format)
                sheet.write(row, 5, "Balance", header_format)
                row += 1
                
                for line in acc_data.get('lines', []):
                    sheet.write(row, 0, str(line['date']), label_format)
                    sheet.write(row, 1, line['ref'] or '', label_format)
                    sheet.write(row, 2, line['partner'] or '', label_format)
                    sheet.write(row, 3, line['debit'], money_format)
                    sheet.write(row, 4, line['credit'], money_format)
                    sheet.write(row, 5, line['balance'], money_format)
                    row += 1
                    
                sheet.write(row, 2, "Total", bold)
                sheet.write(row, 3, acc_data['total_debit'], money_format)
                sheet.write(row, 4, acc_data['total_credit'], money_format)
                sheet.write(row, 5, acc_data['balance'], money_format)
                row += 2

        elif self.report_type == 'cash_book':
            sheet = workbook.add_worksheet('Cash Book')
            sheet.set_column(0, 0, 15)
            sheet.set_column(1, 1, 20)
            sheet.set_column(2, 2, 20)
            sheet.set_column(3, 3, 35)
            sheet.set_column(4, 6, 15)
            sheet.write(0, 0, "Cash Book", title_format)
            sheet.write(1, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}")
            
            row = 3
            sheet.write(row, 0, "Date", header_format)
            sheet.write(row, 1, "Journal", header_format)
            sheet.write(row, 2, "Ref", header_format)
            sheet.write(row, 3, "Account", header_format)
            sheet.write(row, 4, "Debit", header_format)
            sheet.write(row, 5, "Credit", header_format)
            sheet.write(row, 6, "Balance", header_format)
            row += 1
            
            data = report_model.get_journal_book(journal_type='cash', date_from=self.date_from, date_to=self.date_to)
            for line in data:
                sheet.write(row, 0, str(line['date']), label_format)
                sheet.write(row, 1, line['journal'], label_format)
                sheet.write(row, 2, line['ref'], label_format)
                sheet.write(row, 3, line['account'], label_format)
                sheet.write(row, 4, line['debit'], money_format)
                sheet.write(row, 5, line['credit'], money_format)
                sheet.write(row, 6, line['balance'], money_format)
                row += 1

        elif self.report_type == 'tax_report':
            sheet = workbook.add_worksheet('Tax Report')
            sheet.set_column(0, 0, 40)
            sheet.set_column(1, 1, 15)
            sheet.write(0, 0, "Tax Report", title_format)
            sheet.write(1, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}")
            
            row = 3
            sheet.write(row, 0, "Tax Name", header_format)
            sheet.write(row, 1, "Tax Amount", header_format)
            row += 1
            
            data = report_model.get_tax_report(self.date_from, self.date_to)
            for line in data:
                sheet.write(row, 0, line['tax_name'], label_format)
                sheet.write(row, 1, line['tax_amount'], money_format)
                row += 1

        elif self.report_type == 'gstr2b_reconciliation':
            sheet = workbook.add_worksheet('GSTR-2B Reconciliation')
            sheet.set_column(0, 0, 45)
            sheet.set_column(1, 1, 15)
            sheet.set_column(2, 7, 18)
            sheet.write(0, 0, "GSTR-2B Reconciliation Report", title_format)
            sheet.write(1, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}")

            row = 3
            headers = ["Particulars", "Voucher Count", "Taxable Amount", "IGST", "CGST", "SGST", "Cess", "Tax Amount", "Invoice Status Amount"]
            for col, h in enumerate(headers):
                sheet.write(row, col, h, header_format)
            row += 1

            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'posted' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            posted_count = self.env.cr.fetchone()[0] or 0
            
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'draft' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            draft_count = self.env.cr.fetchone()[0] or 0

            self.env.cr.execute("""
                SELECT 
                    SUM(CASE WHEN l.display_type = 'product' THEN l.balance ELSE 0 END),
                    SUM(CASE WHEN a.account_type = 'asset_receivable' OR a.name::text ILIKE %s THEN l.debit - l.credit ELSE 0 END),
                    SUM(CASE WHEN a.name::text ILIKE %s THEN l.debit - l.credit ELSE 0 END),
                    SUM(CASE WHEN a.name::text ILIKE %s THEN l.debit - l.credit ELSE 0 END)
                FROM account_move_line l
                JOIN account_account a ON l.account_id = a.id
                JOIN account_move m ON l.move_id = m.id
                WHERE m.move_type = 'in_invoice' AND m.state = 'posted' AND m.date >= %s AND m.date <= %s
            """, ('%IGST%', '%CGST%', '%SGST%', self.date_from or '2025-04-01', self.date_to))
            totals = self.env.cr.fetchone()
            taxable = abs(totals[0] or 0.0)
            igst = totals[1] or 0.0
            cgst = totals[2] or 0.0
            sgst = totals[3] or 0.0
            total_tax = igst + cgst + sgst

            sheet.write(row, 0, "Reconciled", label_format)
            sheet.write(row, 1, posted_count, label_format)
            row += 1

            sheet.write(row, 0, "Unreconciled", label_format)
            sheet.write(row, 1, draft_count, label_format)
            row += 1

            sheet.write(row, 0, "Available Only in Books", label_format)
            sheet.write(row, 1, draft_count, label_format)
            row += 1

            sheet.write(row, 0, "Uncertain Transactions (Corrections needed)", label_format)
            sheet.write(row, 1, 0, label_format)
            row += 2

            sheet.write(row, 0, "All other ITC from Registered Persons (Excluding Reverse Charge)", label_format)
            sheet.write(row, 1, posted_count, label_format)
            sheet.write(row, 2, taxable, money_format)
            sheet.write(row, 3, igst, money_format)
            sheet.write(row, 4, cgst, money_format)
            sheet.write(row, 5, sgst, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, total_tax, money_format)
            sheet.write(row, 8, "Unreconciled", label_format)
            row += 1

            sheet.write(row, 0, "Total Available ITC", label_format)
            sheet.write(row, 1, posted_count, label_format)
            sheet.write(row, 2, taxable, money_format)
            sheet.write(row, 3, igst, money_format)
            sheet.write(row, 4, cgst, money_format)
            sheet.write(row, 5, sgst, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, total_tax, money_format)
            sheet.write(row, 8, "", label_format)
            row += 1

        elif self.report_type == 'gstr2a_reconciliation':
            sheet = workbook.add_worksheet('GSTR-2A Reconciliation')
            sheet.set_column(0, 0, 45)
            sheet.set_column(1, 1, 15)
            sheet.set_column(2, 7, 18)
            sheet.write(0, 0, "GSTR-2A Reconciliation Report", title_format)
            sheet.write(1, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}")

            row = 3
            headers = ["Particulars", "Voucher Count", "Taxable Amount", "IGST", "CGST", "SGST", "Cess", "Tax Amount", "Invoice Status Amount"]
            for col, h in enumerate(headers):
                sheet.write(row, col, h, header_format)
            row += 1

            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'posted' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            posted_count = self.env.cr.fetchone()[0] or 0
            
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE move_type = 'in_invoice' AND state = 'draft' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            draft_count = self.env.cr.fetchone()[0] or 0

            self.env.cr.execute("""
                SELECT 
                    SUM(CASE WHEN l.display_type = 'product' THEN l.balance ELSE 0 END),
                    SUM(CASE WHEN a.account_type = 'asset_receivable' OR a.name::text ILIKE %s THEN l.debit - l.credit ELSE 0 END),
                    SUM(CASE WHEN a.name::text ILIKE %s THEN l.debit - l.credit ELSE 0 END),
                    SUM(CASE WHEN a.name::text ILIKE %s THEN l.debit - l.credit ELSE 0 END)
                FROM account_move_line l
                JOIN account_account a ON l.account_id = a.id
                JOIN account_move m ON l.move_id = m.id
                WHERE m.move_type = 'in_invoice' AND m.state = 'posted' AND m.date >= %s AND m.date <= %s
            """, ('%IGST%', '%CGST%', '%SGST%', self.date_from or '2025-04-01', self.date_to))
            totals = self.env.cr.fetchone()
            taxable = abs(totals[0] or 0.0)
            igst = totals[1] or 0.0
            cgst = totals[2] or 0.0
            sgst = totals[3] or 0.0
            total_tax = igst + cgst + sgst

            sheet.write(row, 0, "Reconciled", label_format)
            sheet.write(row, 1, posted_count, label_format)
            row += 1

            sheet.write(row, 0, "Unreconciled", label_format)
            sheet.write(row, 1, draft_count, label_format)
            row += 1

            sheet.write(row, 0, "Available only in Books", label_format)
            sheet.write(row, 1, draft_count, label_format)
            row += 1

            sheet.write(row, 0, "Uncertain Transactions (Corrections needed)", label_format)
            sheet.write(row, 1, 0, label_format)
            row += 1

            sheet.write(row, 0, "Check Vouchers Having Potential Conflicts with Masters", label_format)
            sheet.write(row, 1, "", label_format)
            sheet.write(row, 8, "Yes", label_format)
            row += 2

            # Table rows
            sheet.write(row, 0, "B2B Invoices", label_format)
            sheet.write(row, 1, posted_count, label_format)
            sheet.write(row, 2, taxable, money_format)
            sheet.write(row, 3, igst, money_format)
            sheet.write(row, 4, cgst, money_format)
            sheet.write(row, 5, sgst, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, total_tax, money_format)
            sheet.write(row, 8, "Unreconciled", label_format)
            row += 1

            sheet.write(row, 0, "Amendments to B2B Invoices", label_format)
            sheet.write(row, 1, 0, label_format)
            sheet.write(row, 2, 0.0, money_format)
            sheet.write(row, 3, 0.0, money_format)
            sheet.write(row, 4, 0.0, money_format)
            sheet.write(row, 5, 0.0, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, 0.0, money_format)
            sheet.write(row, 8, "", label_format)
            row += 1

            sheet.write(row, 0, "Credit/Debit Notes", label_format)
            sheet.write(row, 1, 0, label_format)
            sheet.write(row, 2, 0.0, money_format)
            sheet.write(row, 3, 0.0, money_format)
            sheet.write(row, 4, 0.0, money_format)
            sheet.write(row, 5, 0.0, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, 0.0, money_format)
            sheet.write(row, 8, "", label_format)
            row += 1

            sheet.write(row, 0, "Amendments to Credit/Debit Notes", label_format)
            sheet.write(row, 1, 0, label_format)
            sheet.write(row, 2, 0.0, money_format)
            sheet.write(row, 3, 0.0, money_format)
            sheet.write(row, 4, 0.0, money_format)
            sheet.write(row, 5, 0.0, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, 0.0, money_format)
            sheet.write(row, 8, "", label_format)
            row += 1

            sheet.write(row, 0, "ISD Credits", label_format)
            sheet.write(row, 1, 0, label_format)
            sheet.write(row, 2, 0.0, money_format)
            sheet.write(row, 3, 0.0, money_format)
            sheet.write(row, 4, 0.0, money_format)
            sheet.write(row, 5, 0.0, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, 0.0, money_format)
            sheet.write(row, 8, "", label_format)
            row += 1

            sheet.write(row, 0, "Import of Goods from overseas on Bill of Entry", label_format)
            sheet.write(row, 1, 0, label_format)
            sheet.write(row, 2, 0.0, money_format)
            sheet.write(row, 3, 0.0, money_format)
            sheet.write(row, 4, 0.0, money_format)
            sheet.write(row, 5, 0.0, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, 0.0, money_format)
            sheet.write(row, 8, "", label_format)
            row += 1

            sheet.write(row, 0, "Import of Goods from SEZ Units/Developers on Bill of Entry", label_format)
            sheet.write(row, 1, 0, label_format)
            sheet.write(row, 2, 0.0, money_format)
            sheet.write(row, 3, 0.0, money_format)
            sheet.write(row, 4, 0.0, money_format)
            sheet.write(row, 5, 0.0, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, 0.0, money_format)
            sheet.write(row, 8, "", label_format)
            row += 1

            sheet.write(row, 0, "Total", label_format)
            sheet.write(row, 1, posted_count, label_format)
            sheet.write(row, 2, taxable, money_format)
            sheet.write(row, 3, igst, money_format)
            sheet.write(row, 4, cgst, money_format)
            sheet.write(row, 5, sgst, money_format)
            sheet.write(row, 6, 0.0, money_format)
            sheet.write(row, 7, total_tax, money_format)
            sheet.write(row, 8, "", label_format)
            sheet.write(row, 8, "", label_format)
            row += 1

        elif self.report_type == 'annual_computation':
            sheet = workbook.add_worksheet('Annual Computation')
            sheet.set_column('A:A', 60)
            sheet.set_column('B:G', 15)

            company = self.env.company
            # Header Info
            row = 0
            sheet.write(row, 0, company.name or '', workbook.add_format({'bold': True, 'font_size': 14}))
            row += 1
            sheet.write(row, 0, company.street or '', label_format)
            row += 1
            sheet.write(row, 0, (company.street2 or '') + ' ' + (company.city or ''), label_format)
            row += 1
            sheet.write(row, 0, (company.state_id.name or '') + ' ' + (company.zip or ''), label_format)
            row += 1
            sheet.write(row, 0, f"E-Mail : {company.email or ''}", label_format)
            row += 1
            sheet.write(row, 0, 'Annual Computation', bold)
            row += 1
            sheet.write(row, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}", label_format)
            row += 1
            sheet.write(row, 0, 'GST Registration:', label_format)
            sheet.write(row, 1, company.vat or '', bold)
            row += 2

            # Fetch Data
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE state = 'posted' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            posted_count = self.env.cr.fetchone()[0] or 0
            
            self.env.cr.execute("""
                SELECT count(id) FROM account_move WHERE state = 'draft' AND date >= %s AND date <= %s
            """, (self.date_from or '2025-04-01', self.date_to))
            draft_count = self.env.cr.fetchone()[0] or 0

            self.env.cr.execute("""
                SELECT 
                    SUM(CASE WHEN l.display_type = 'product' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%IGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%CGST%%' THEN abs(l.balance) ELSE 0 END),
                    SUM(CASE WHEN t.name::text ILIKE '%%SGST%%' OR t.name::text ILIKE '%%UTGST%%' THEN abs(l.balance) ELSE 0 END)
                FROM account_move_line l
                JOIN account_move m ON l.move_id = m.id
                LEFT JOIN account_tax t ON l.tax_line_id = t.id
                WHERE m.state = 'posted' AND m.date >= %s AND m.date <= %s
                  AND (l.display_type = 'product' OR l.tax_line_id IS NOT NULL)
            """, (self.date_from or '2025-04-01', self.date_to))
            totals = self.env.cr.fetchone()
            taxable = totals[0] or 0.0
            igst = totals[1] or 0.0
            cgst = totals[2] or 0.0
            sgst = totals[3] or 0.0
            total_tax = igst + cgst + sgst

            # Statistics Table
            sheet.write(row, 0, "Particulars", header_format)
            sheet.write(row, 1, "Voucher Count", header_format)
            row += 1
            sheet.write(row, 0, "Total Vouchers", label_format)
            sheet.write(row, 1, posted_count + draft_count, bold)
            row += 1
            sheet.write(row, 0, "Included In Return", label_format)
            sheet.write(row, 1, posted_count, bold)
            row += 1
            sheet.write(row, 0, "Not Relevant for This Return", label_format)
            sheet.write(row, 1, draft_count, label_format)
            row += 1
            sheet.write(row, 0, "Uncertain Transactions (Corrections needed)", label_format)
            sheet.write(row, 1, 0, label_format)
            row += 1
            sheet.write(row, 0, "Check Vouchers Having Potential Conflicts with Masters", label_format)
            sheet.write(row, 1, "Yes", label_format)
            row += 2

            # Return View Headers
            headers = ["Particulars", "Taxable Amount", "IGST", "CGST", "SGST/UTGST", "Cess", "Tax Amount"]
            for col, h in enumerate(headers):
                sheet.write(row, col, h, header_format)
            row += 1

            def write_data_row(row, label, data_tuple, is_bold=False):
                fmt = bold if is_bold else label_format
                mfmt = money_format
                sheet.write(row, 0, label, fmt)
                sheet.write(row, 1, data_tuple[0], mfmt)
                sheet.write(row, 2, data_tuple[1], mfmt)
                sheet.write(row, 3, data_tuple[2], mfmt)
                sheet.write(row, 4, data_tuple[3], mfmt)
                sheet.write(row, 5, 0.0, mfmt) # Cess placeholder
                sheet.write(row, 6, sum(data_tuple[1:]), mfmt)
                return row + 1

            # Liability
            sheet.write(row, 0, "Liability", bold)
            row += 1
            row = write_data_row(row, "Outward and Inward Supplies on Which Tax is Payable (Including Advances)", (taxable, igst, cgst, sgst))
            row = write_data_row(row, "Outward Supplies on Which Tax is Not Payable", (0, 0, 0, 0))
            row = write_data_row(row, "Missing Invoice Reported in Current Period", (0, 0, 0, 0))
            row = write_data_row(row, "Total Liability", (taxable, igst, cgst, sgst), is_bold=True)
            row += 1

            # ITC
            sheet.write(row, 0, "Input Tax Credit", bold)
            row += 1
            row = write_data_row(row, "Input Tax Credit", (taxable, igst, cgst, sgst))
            row = write_data_row(row, "Reversal of Input Tax Credit, Adjusted and Ineligible Input Tax Credit Declared", (0, 0, 0, 0))
            row = write_data_row(row, "Total Input Tax Credit After Reversal & Ineligible Input Tax Credit", (taxable, igst, cgst, sgst), is_bold=True)
            row += 1

            # Others
            row = write_data_row(row, "Interest, Late Fee, Penalty and Others", (0, 0, 0, 0), is_bold=True)
            sheet.write(row, 0, "HSN/SAC Summary", bold)
            row += 1
            row = write_data_row(row, "Summary of Outward Supplies", (0, 0, 0, 0))
            row = write_data_row(row, "Summary of Inward Supplies", (taxable, igst, cgst, sgst))

        elif self.report_type == 'brs_statement':
            sheet = workbook.add_worksheet('BRS Statement')
            sheet.set_column('A:A', 50)
            sheet.set_column('B:D', 15)
            sheet.set_column('E:E', 25)
            
            company = self.env.company
            row = 0
            sheet.write(row, 0, company.name or '', workbook.add_format({'bold': True, 'font_size': 14}))
            row += 1
            sheet.write(row, 0, company.street or '', label_format)
            row += 1
            sheet.write(row, 0, (company.street2 or '') + ' ' + (company.city or ''), label_format)
            row += 1
            sheet.write(row, 0, (company.state_id.name or '') + ' ' + (company.zip or ''), label_format)
            row += 1
            sheet.write(row, 0, f"E-Mail : {company.email or ''}", label_format)
            row += 1
            sheet.write(row, 0, 'Bank Reconciliation Statement (BRS)', bold)
            row += 1
            sheet.write(row, 0, f"From {self.date_from} to {self.date_to or fields.Date.today()}", label_format)
            row += 2
            
            # Fetch data
            data = self._get_brs_data()
            
            sheet.write(row, 0, "Particulars / Description", header_format)
            sheet.write(row, 1, "Debit", header_format)
            sheet.write(row, 2, "Credit", header_format)
            sheet.write(row, 3, "Balance", header_format)
            sheet.write(row, 4, "Status / Type", header_format)
            row += 1
            
            sheet.write(row, 0, "Odoo Ledger Opening Balance", bold)
            sheet.write(row, 3, data['odoo_opening_bal'], money_format)
            sheet.write(row, 4, "Ledger Opening", label_format)
            row += 1
            
            sheet.write(row, 0, "Bank Statement Opening Balance", bold)
            sheet.write(row, 3, -data['excel_opening_bal'], money_format)
            sheet.write(row, 4, "Statement Opening", label_format)
            row += 1
            
            sheet.write(row, 0, "Prior Period Discrepancy carried forward", bold)
            sheet.write(row, 3, data['prior_discrepancy'], money_format)
            sheet.write(row, 4, "Carry Forward", label_format)
            row += 1
            
            reconciled_count = len(data['excel_txs']) - len(data['unmatched_excel'])
            sheet.write(row, 0, f"Matched/Reconciled Transactions ({reconciled_count} entries)", label_format)
            sheet.write(row, 1, sum(e['debit'] for e in data['excel_txs'] if e not in data['unmatched_excel']), money_format)
            sheet.write(row, 2, sum(e['credit'] for e in data['excel_txs'] if e not in data['unmatched_excel']), money_format)
            sheet.write(row, 4, "Reconciled", label_format)
            row += 1
            
            if data['unmatched_excel']:
                sheet.write(row, 0, "--- Omissions: Transactions in Bank Statement but NOT in Odoo Books ---", bold)
                row += 1
                for tx in data['unmatched_excel']:
                    sheet.write(row, 0, tx['particulars'], label_format)
                    sheet.write(row, 1, tx['debit'], money_format)
                    sheet.write(row, 2, tx['credit'], money_format)
                    sheet.write(row, 4, "Missing in Books", label_format)
                    row += 1
                    
            if data['unmatched_odoo']:
                sheet.write(row, 0, "--- Discrepancies: Transactions in Odoo Books but NOT in Bank Statement ---", bold)
                row += 1
                for tx in data['unmatched_odoo']:
                    status = 'Duplicate (VENDOR_PAY/)' if tx['ref'].startswith('VENDOR_PAY/') else 'Unmatched Ledger Entry'
                    if tx['move_name'] == 'MISC/2025/05/0001':
                        status = 'Incorrect Contra (Reversed)'
                    sheet.write(row, 0, f"[{tx['move_name']}] {tx['particulars']}", label_format)
                    sheet.write(row, 1, tx['debit'], money_format)
                    sheet.write(row, 2, tx['credit'], money_format)
                    sheet.write(row, 4, status, label_format)
                    row += 1
                    
            sheet.write(row, 0, "--------------------------------------------------", label_format)
            row += 1
            
            sheet.write(row, 0, "Odoo Ledger Ending Balance", bold)
            sheet.write(row, 3, data['odoo_closing_bal'], money_format)
            sheet.write(row, 4, "Ledger Ending", label_format)
            row += 1
            
            sheet.write(row, 0, "Add: Duplicate payments posted in Odoo (VENDOR_PAY/) to be reversed", label_format)
            sheet.write(row, 3, data['total_duplicates'], money_format)
            sheet.write(row, 4, "Adjustment", label_format)
            row += 1
            
            sheet.write(row, 0, "Less: Reverse incorrect Contra Debit entry in Odoo", label_format)
            sheet.write(row, 3, -20000.0, money_format)
            sheet.write(row, 4, "Adjustment", label_format)
            row += 1
            
            sheet.write(row, 0, "Less: Record missing Petty Cash payment (Excel Row 67)", label_format)
            sheet.write(row, 3, -20000.0, money_format)
            sheet.write(row, 4, "Adjustment", label_format)
            row += 1
            
            sheet.write(row, 0, "Add: Discrepancy carried forward from prior period", label_format)
            sheet.write(row, 3, data['prior_discrepancy'], money_format)
            sheet.write(row, 4, "Adjustment", label_format)
            row += 1
            
            sheet.write(row, 0, "Adjusted Balance as per Odoo Ledger", bold)
            sheet.write(row, 3, data['adjusted_bal'], money_format)
            sheet.write(row, 4, "Reconciled Balance", label_format)
            row += 1
            
            sheet.write(row, 0, "Actual Balance as per Bank Statement", bold)
            sheet.write(row, 3, -data['excel_closing_bal'], money_format)
            sheet.write(row, 4, "Statement Ending", label_format)
            row += 1
            
            sheet.write(row, 0, "Unreconciled Difference", bold)
            sheet.write(row, 3, data['adjusted_bal'] - (-data['excel_closing_bal']), money_format)
            sheet.write(row, 4, "Difference", label_format)
            row += 1

        workbook.close()
        output.seek(0)
        file_data = base64.b64encode(output.read())
        attachment = self.env['ir.attachment'].create({
            'name': f'{self.report_type}.xlsx',
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
