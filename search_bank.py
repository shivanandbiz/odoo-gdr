import openpyxl

wb = openpyxl.load_workbook('/home/biz/odoo/all_tally_to_odoo_migratation.xlsx', read_only=True, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        for c_idx, cell in enumerate(row):
            if cell and "HDFC" in str(cell):
                print(f"Sheet: {sheet_name}, Row: {r_idx+1}, Col: {c_idx+1}, Value: {cell}")
